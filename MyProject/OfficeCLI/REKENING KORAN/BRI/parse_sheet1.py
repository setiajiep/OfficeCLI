import openpyxl

wb = openpyxl.load_workbook('Akr_Payment.xlsx', data_only=True)
ws = wb['Sheet1']

with open('sheet1_dump.txt', 'w') as f:
    f.write(f"Max row: {ws.max_row}\n")
    for r in range(1, ws.max_row + 1):
        cols_val = {}
        for c in range(1, 35):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                cols_val[c] = v
        if cols_val:
            f.write(f"Row {r:3d}: {cols_val}\n")

print("Dumped to sheet1_dump.txt")
