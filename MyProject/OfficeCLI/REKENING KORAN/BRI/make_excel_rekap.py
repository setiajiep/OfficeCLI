import glob
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Load all CSV data
files = sorted(glob.glob('*.csv'))
dfs = []
for f in files:
    m = f.split('_')[0]
    df = pd.read_csv(f)
    df['BULAN'] = m
    dfs.append(df)

df_all = pd.concat(dfs, ignore_index=True)
df_all['MUTASI_DEBET'] = df_all['MUTASI_DEBET'].astype(float)
df_all['MUTASI_KREDIT'] = df_all['MUTASI_KREDIT'].astype(float)
df_all['SALDO_AWAL_MUTASI'] = df_all['SALDO_AWAL_MUTASI'].astype(float)
df_all['SALDO_AKHIR_MUTASI'] = df_all['SALDO_AKHIR_MUTASI'].astype(float)

# Categorization logic
def categorize_refined(row):
    desk = str(row['DESK_TRAN']).upper()
    rem = str(row['REMARK_CUSTOM']).upper()
    gl = str(row['GLSIGN']).upper()
    deb = row['MUTASI_DEBET']
    kre = row['MUTASI_KREDIT']

    if gl == 'CR' or kre > 0:
        if 'INTEREST' in desk or 'INTEREST' in rem:
            return 'Pendapatan Bunga Bank'
        elif 'AKR' in desk or 'AKR' in rem:
            return 'Penerimaan Penjualan (AKR Corporindo)'
        elif 'TUNAI' in desk or 'TUNAI' in rem:
            return 'Setoran Tunai'
        elif 'PNCAIRN' in rem or 'PEMINDAHAN BUKUAN AN PT HIJAU' in rem:
            return 'Pencairan Kredit & Pemindahbukuan Internal'
        elif 'TRANSFER BI-FAST DARI' in rem or 'BFST' in desk or 'BFST' in rem or 'NBMB' in rem or 'TRANSFER DARI' in rem:
            return 'Transfer Masuk BI-Fast / Pihak Ketiga'
        else:
            return 'Transfer Masuk Lainnya'
    else:
        if 'TAX' in desk or 'TAX' in rem:
            return 'Pajak Bunga Tabungan'
        elif 'PROVISI' in rem or 'ADM KMK' in rem or 'ASURANSI' in rem or 'NOTARIS' in rem:
            return 'Biaya Provisi, Notaris & Asuransi Kredit'
        elif 'MONTHLY FEE' in desk or 'MINIMUM BALANCE' in desk or deb == 6500.0 or (deb in [1500.0, 3000.0] and ('PLN' in rem or 'PUL-' in rem)):
            return 'Biaya Admin Bank & Transfer'
        elif 'PLN' in rem or 'PUL-' in rem:
            return 'Pembayaran Tagihan Listrik & Pulsa'
        elif 'BPJS' in desk or 'BPJS' in rem:
            return 'Pembayaran BPJS'
        elif 'BRIVA' in rem or 'DANA' in rem:
            return 'Pembayaran e-Wallet (BRIVA/DANA)'
        elif 'DEPLESI' in rem or 'BILL PAYMENT' in desk or 'BILL PAYMENT' in rem or 'ANG' in rem or 'PNJ' in rem or 'PDBTAN' in rem or 'OB KWJIBN' in rem:
            return 'Pembayaran Angsuran, Pinjaman & Deplesi'
        elif 'FROM:' in desk or 'FROM:' in rem:
            return 'Transfer Antar Rekening (Overbooking)'
        elif 'TRANSFER BI-FAST KE' in rem or 'BFST' in desk or 'ATMSTRPRM' in desk or 'TO SARI WINDARSIH' in rem or 'ESB:INDS:' in desk:
            return 'Transfer Keluar BI-Fast / Bank Transfer'
        else:
            return 'Pengeluaran Lainnya'

df_all['KATEGORI'] = df_all.apply(categorize_refined, axis=1)

# Create Workbook
wb = openpyxl.Workbook()
# remove default sheet
wb.remove(wb.active)

# Styling Definitions
FONT_NAME = 'Segoe UI'

font_title = Font(name=FONT_NAME, size=16, bold=True, color='1F4E78')
font_subtitle = Font(name=FONT_NAME, size=10, italic=True, color='595959')
font_section = Font(name=FONT_NAME, size=12, bold=True, color='1F4E78')
font_header = Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')
font_card_num = Font(name=FONT_NAME, size=14, bold=True, color='1F4E78')
font_card_lbl = Font(name=FONT_NAME, size=9, bold=True, color='595959')

font_bold = Font(name=FONT_NAME, size=10, bold=True)
font_regular = Font(name=FONT_NAME, size=10)
font_kredit = Font(name=FONT_NAME, size=10, color='2E75B6')
font_debet = Font(name=FONT_NAME, size=10, color='C00000')

fill_navy = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
fill_blue_sub = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
fill_soft_blue = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
fill_card = PatternFill(start_color='F2F4F8', end_color='F2F4F8', fill_type='solid')
fill_zebra = PatternFill(start_color='F9FAFC', end_color='F9FAFC', fill_type='solid')
fill_green_tint = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
fill_red_tint = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
fill_total = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

thin_border = Side(border_style='thin', color='D9D9D9')
thick_bottom = Side(border_style='medium', color='1F4E78')
double_bottom = Side(border_style='double', color='1F4E78')
top_thin = Side(border_style='thin', color='1F4E78')

border_all_thin = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
border_header = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thick_bottom)
border_total = Border(top=top_thin, bottom=double_bottom, left=thin_border, right=thin_border)
border_card = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

CURRENCY_FORMAT = 'Rp #,##0.00;[Red](Rp #,##0.00);"-"'
NUMBER_FORMAT = '#,##0'

# ==============================================================================
# SHEET 1: RINGKASAN EKSEKUTIF
# ==============================================================================
ws1 = wb.create_sheet(title='Ringkasan Eksekutif')
ws1.views.sheetView[0].showGridLines = True

# Title Block
ws1['A1'] = "REKAPITULASI REKENING KORAN BANK BRI"
ws1['A1'].font = font_title
ws1['A2'] = "No. Rekening: 13001002256309 | Periode: Januari - Juli 2026 | Mata Uang: IDR"
ws1['A2'].font = font_subtitle

# KPI Cards
# Card 1: Saldo Awal
ws1.merge_cells('A4:B4')
ws1['A4'] = "SALDO AWAL (05/01/2026)"
ws1['A4'].font = font_card_lbl
ws1['A4'].alignment = align_center
ws1['A4'].fill = fill_card
ws1.merge_cells('A5:B5')
ws1['A5'] = df_all['SALDO_AWAL_MUTASI'].iloc[0]
ws1['A5'].font = font_card_num
ws1['A5'].number_format = CURRENCY_FORMAT
ws1['A5'].alignment = align_center
ws1['A5'].fill = fill_card

# Card 2: Total Uang Masuk
ws1.merge_cells('C4:D4')
ws1['C4'] = "TOTAL UANG MASUK (KREDIT)"
ws1['C4'].font = font_card_lbl
ws1['C4'].alignment = align_center
ws1['C4'].fill = fill_card
ws1.merge_cells('C5:D5')
ws1['C5'] = df_all['MUTASI_KREDIT'].sum()
ws1['C5'].font = Font(name=FONT_NAME, size=14, bold=True, color='2E75B6')
ws1['C5'].number_format = CURRENCY_FORMAT
ws1['C5'].alignment = align_center
ws1['C5'].fill = fill_card

# Card 3: Total Uang Keluar
ws1.merge_cells('E4:F4')
ws1['E4'] = "TOTAL UANG KELUAR (DEBET)"
ws1['E4'].font = font_card_lbl
ws1['E4'].alignment = align_center
ws1['E4'].fill = fill_card
ws1.merge_cells('E5:F5')
ws1['E5'] = df_all['MUTASI_DEBET'].sum()
ws1['E5'].font = Font(name=FONT_NAME, size=14, bold=True, color='C00000')
ws1['E5'].number_format = CURRENCY_FORMAT
ws1['E5'].alignment = align_center
ws1['E5'].fill = fill_card

# Card 4: Net Cash Flow
ws1.merge_cells('G4:H4')
ws1['G4'] = "NET CASH FLOW"
ws1['G4'].font = font_card_lbl
ws1['G4'].alignment = align_center
ws1['G4'].fill = fill_card
ws1.merge_cells('G5:H5')
ws1['G5'] = "=C5-E5"
ws1['G5'].font = font_card_num
ws1['G5'].number_format = CURRENCY_FORMAT
ws1['G5'].alignment = align_center
ws1['G5'].fill = fill_card

# Card 5: Saldo Akhir
ws1.merge_cells('I4:J4')
ws1['I4'] = "SALDO AKHIR (30/07/2026)"
ws1['I4'].font = font_card_lbl
ws1['I4'].alignment = align_center
ws1['I4'].fill = fill_card
ws1.merge_cells('I5:J5')
ws1['I5'] = df_all['SALDO_AKHIR_MUTASI'].iloc[-1]
ws1['I5'].font = font_card_num
ws1['I5'].number_format = CURRENCY_FORMAT
ws1['I5'].alignment = align_center
ws1['I5'].fill = fill_card

for row in range(4, 6):
    for col in range(1, 11):
        cell = ws1.cell(row=row, column=col)
        cell.border = border_card

# Section 1: Tabel Ringkasan Bulanan
ws1['A7'] = "1. RINGKASAN ARUS KAS BULANAN (JANUARI - JULI 2026)"
ws1['A7'].font = font_section

headers_m = ['Periode Bulan', 'Saldo Awal (Rp)', 'Total Masuk / Kredit (Rp)', 'Total Keluar / Debet (Rp)', 'Net Cash Flow (Rp)', 'Saldo Akhir (Rp)', 'Jumlah Transaksi']
for col_num, h_text in enumerate(headers_m, 1):
    cell = ws1.cell(row=8, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

monthly_summary = []
for f in files:
    m = f.split('_')[0]
    df_m = df_all[df_all['BULAN'] == m]
    s_awal = df_m['SALDO_AWAL_MUTASI'].iloc[0]
    s_akhir = df_m['SALDO_AKHIR_MUTASI'].iloc[-1]
    monthly_summary.append((m, s_awal, s_akhir))

row_start = 9
for idx, (m, s_awal, s_akhir) in enumerate(monthly_summary):
    r = row_start + idx
    # Month name formatting e.g. Januari 2026
    month_names = {'2026-01':'Januari 2026', '2026-02':'Februari 2026', '2026-03':'Maret 2026', '2026-04':'April 2026', '2026-05':'Mei 2026', '2026-06':'Juni 2026', '2026-07':'Juli 2026'}
    m_label = month_names.get(m, m)
    
    ws1.cell(row=r, column=1, value=m_label).alignment = align_left
    ws1.cell(row=r, column=2, value=s_awal).number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=3, value=f"=SUMIFS('Detail Transaksi'!I:I, 'Detail Transaksi'!B:B, \"{m}\")").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=4, value=f"=SUMIFS('Detail Transaksi'!H:H, 'Detail Transaksi'!B:B, \"{m}\")").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=6, value=s_akhir).number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=7, value=f"=COUNTIFS('Detail Transaksi'!B:B, \"{m}\")").number_format = NUMBER_FORMAT

    # Styling
    fill_cur = fill_zebra if idx % 2 == 1 else PatternFill(fill_type=None)
    for c in range(1, 8):
        cell = ws1.cell(row=r, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_cur.fill_type:
            cell.fill = fill_cur
        if c in [2,3,4,5,6]:
            cell.alignment = align_right
        elif c == 7:
            cell.alignment = align_center

# Total Row for Monthly Summary
r_tot = row_start + len(monthly_summary)
ws1.cell(row=r_tot, column=1, value="TOTAL / SALDO AKHIR").alignment = align_left
ws1.cell(row=r_tot, column=2, value=monthly_summary[0][1]).number_format = CURRENCY_FORMAT # Saldo Awal Jan
ws1.cell(row=r_tot, column=3, value=f"=SUM(C9:C{r_tot-1})").number_format = CURRENCY_FORMAT
ws1.cell(row=r_tot, column=4, value=f"=SUM(D9:D{r_tot-1})").number_format = CURRENCY_FORMAT
ws1.cell(row=r_tot, column=5, value=f"=C{r_tot}-D{r_tot}").number_format = CURRENCY_FORMAT
ws1.cell(row=r_tot, column=6, value=monthly_summary[-1][2]).number_format = CURRENCY_FORMAT # Saldo Akhir Jul
ws1.cell(row=r_tot, column=7, value=f"=SUM(G9:G{r_tot-1})").number_format = NUMBER_FORMAT

for c in range(1, 8):
    cell = ws1.cell(row=r_tot, column=c)
    cell.font = font_bold
    cell.fill = fill_total
    cell.border = border_total
    if c in [2,3,4,5,6]:
        cell.alignment = align_right
    elif c == 7:
        cell.alignment = align_center

# Section 2: Tabel Rekapitulasi Kategori Transaksi
r_cat_head = r_tot + 3
ws1.cell(row=r_cat_head-1, column=1, value="2. REKAPITULASI KATEGORI TRANSAKSI").font = font_section

headers_cat = ['No', 'Kategori Transaksi', 'Jenis Arus Kas', 'Total Debet (Pengeluaran)', 'Total Kredit (Penerimaan)', 'Net Cash Impact', 'Jumlah Transaksi', '% Dari Total Arus']
for col_num, h_text in enumerate(headers_cat, 1):
    cell = ws1.cell(row=r_cat_head, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

# Populate Categories
categories_list = [
    ('Penerimaan Penjualan (AKR Corporindo)', 'Uang Masuk'),
    ('Pencairan Kredit & Pemindahbukuan Internal', 'Uang Masuk'),
    ('Transfer Masuk BI-Fast / Pihak Ketiga', 'Uang Masuk'),
    ('Setoran Tunai', 'Uang Masuk'),
    ('Pendapatan Bunga Bank', 'Uang Masuk'),
    ('Transfer Keluar BI-Fast / Bank Transfer', 'Uang Keluar'),
    ('Pembayaran Angsuran, Pinjaman & Deplesi', 'Uang Keluar'),
    ('Transfer Antar Rekening (Overbooking)', 'Uang Keluar'),
    ('Biaya Provisi, Notaris & Asuransi Kredit', 'Uang Keluar'),
    ('Pembayaran Tagihan Listrik & Pulsa', 'Uang Keluar'),
    ('Pembayaran BPJS', 'Uang Keluar'),
    ('Pembayaran e-Wallet (BRIVA/DANA)', 'Uang Keluar'),
    ('Biaya Admin Bank & Transfer', 'Uang Keluar'),
    ('Pajak Bunga Tabungan', 'Uang Keluar')
]

r_cat_start = r_cat_head + 1
for idx, (cat_name, cat_type) in enumerate(categories_list, 1):
    r = r_cat_start + idx - 1
    ws1.cell(row=r, column=1, value=idx).alignment = align_center
    ws1.cell(row=r, column=2, value=cat_name).alignment = align_left
    ws1.cell(row=r, column=3, value=cat_type).alignment = align_center
    ws1.cell(row=r, column=4, value=f"=SUMIFS('Detail Transaksi'!H:H, 'Detail Transaksi'!E:E, B{r})").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=5, value=f"=SUMIFS('Detail Transaksi'!I:I, 'Detail Transaksi'!E:E, B{r})").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=6, value=f"=E{r}-D{r}").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=7, value=f"=COUNTIFS('Detail Transaksi'!E:E, B{r})").number_format = NUMBER_FORMAT
    ws1.cell(row=r, column=8, value=f"=(D{r}+E{r})/('Ringkasan Eksekutif'!C5+'Ringkasan Eksekutif'!E5)").number_format = '0.00%'

    fill_cur = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
    for c in range(1, 9):
        cell = ws1.cell(row=r, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_cur.fill_type:
            cell.fill = fill_cur
        if c in [4, 5, 6, 8]:
            cell.alignment = align_right

# Category Total Row
r_cat_tot = r_cat_start + len(categories_list)
ws1.cell(row=r_cat_tot, column=1, value="").alignment = align_center
ws1.cell(row=r_cat_tot, column=2, value="TOTAL KATEGORI").alignment = align_left
ws1.cell(row=r_cat_tot, column=3, value="-").alignment = align_center
ws1.cell(row=r_cat_tot, column=4, value=f"=SUM(D{r_cat_start}:D{r_cat_tot-1})").number_format = CURRENCY_FORMAT
ws1.cell(row=r_cat_tot, column=5, value=f"=SUM(E{r_cat_start}:E{r_cat_tot-1})").number_format = CURRENCY_FORMAT
ws1.cell(row=r_cat_tot, column=6, value=f"=E{r_cat_tot}-D{r_cat_tot}").number_format = CURRENCY_FORMAT
ws1.cell(row=r_cat_tot, column=7, value=f"=SUM(G{r_cat_start}:G{r_cat_tot-1})").number_format = NUMBER_FORMAT
ws1.cell(row=r_cat_tot, column=8, value=f"=SUM(H{r_cat_start}:H{r_cat_tot-1})").number_format = '0.00%'

for c in range(1, 9):
    cell = ws1.cell(row=r_cat_tot, column=c)
    cell.font = font_bold
    cell.fill = fill_total
    cell.border = border_total
    if c in [4, 5, 6, 8]:
        cell.alignment = align_right
    elif c == 7:
        cell.alignment = align_center

# ==============================================================================
# SHEET 2: DETAIL TRANSAKSI
# ==============================================================================
ws2 = wb.create_sheet(title='Detail Transaksi')
ws2.views.sheetView[0].showGridLines = True

ws2['A1'] = "DETAIL MUTASI TRANSAKSI REKENING KORAN BRI"
ws2['A1'].font = font_title
ws2['A2'] = "No. Rekening: 13001002256309 | Total Transaksi: 211 Items"
ws2['A2'].font = font_subtitle

headers_det = [
    'No', 'Periode', 'Tanggal & Waktu', 'No. Rekening', 'Kategori',
    'Keterangan / Remark', 'Jenis (GL)', 'Mutasi Debet (Rp)',
    'Mutasi Kredit (Rp)', 'Saldo Akhir (Rp)', 'User ID', 'Kode Tran'
]

for col_num, h_text in enumerate(headers_det, 1):
    cell = ws2.cell(row=4, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

# Populate 211 Rows
for idx, r in df_all.iterrows():
    row_num = idx + 5
    ws2.cell(row=row_num, column=1, value=idx+1).alignment = align_center
    ws2.cell(row=row_num, column=2, value=str(r['BULAN'])).alignment = align_center
    ws2.cell(row=row_num, column=3, value=str(r['TGL_TRAN'])).alignment = align_center
    ws2.cell(row=row_num, column=4, value=str(r['NOREK'])).alignment = align_center
    ws2.cell(row=row_num, column=5, value=str(r['KATEGORI'])).alignment = align_left
    
    # Clean description / remark
    rem_text = str(r['REMARK_CUSTOM']) if pd.notna(r['REMARK_CUSTOM']) and str(r['REMARK_CUSTOM']).strip() != "" else str(r['DESK_TRAN'])
    ws2.cell(row=row_num, column=6, value=rem_text).alignment = align_left
    
    gl_type = 'Kredit' if r['GLSIGN'] == 'Cr' or r['MUTASI_KREDIT'] > 0 else 'Debet'
    ws2.cell(row=row_num, column=7, value=gl_type).alignment = align_center
    
    ws2.cell(row=row_num, column=8, value=r['MUTASI_DEBET']).number_format = CURRENCY_FORMAT
    ws2.cell(row=row_num, column=9, value=r['MUTASI_KREDIT']).number_format = CURRENCY_FORMAT
    ws2.cell(row=row_num, column=10, value=r['SALDO_AKHIR_MUTASI']).number_format = CURRENCY_FORMAT
    ws2.cell(row=row_num, column=11, value=str(r['TRUSER']) if pd.notna(r['TRUSER']) else '-').alignment = align_center
    ws2.cell(row=row_num, column=12, value=str(r['KODE_TRAN']) if pd.notna(r['KODE_TRAN']) else '-').alignment = align_center

    # Row borders & tinting
    fill_row = fill_zebra if idx % 2 == 1 else PatternFill(fill_type=None)
    if gl_type == 'Kredit' and r['MUTASI_KREDIT'] >= 50000000:
        fill_row = fill_green_tint
    elif gl_type == 'Debet' and r['MUTASI_DEBET'] >= 100000000:
        fill_row = fill_red_tint

    for c in range(1, 13):
        cell = ws2.cell(row=row_num, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_row.fill_type:
            cell.fill = fill_row
        if c in [8, 9, 10]:
            cell.alignment = align_right
            if c == 9 and r['MUTASI_KREDIT'] > 0:
                cell.font = font_kredit
            elif c == 8 and r['MUTASI_DEBET'] > 0:
                cell.font = font_debet

# Total Row for Detail
r_det_tot = len(df_all) + 5
ws2.cell(row=r_det_tot, column=1, value="").alignment = align_center
ws2.cell(row=r_det_tot, column=2, value="TOTAL").alignment = align_center
for c in range(3, 8):
    ws2.cell(row=r_det_tot, column=c, value="-").alignment = align_center

ws2.cell(row=r_det_tot, column=8, value=f"=SUM(H5:H{r_det_tot-1})").number_format = CURRENCY_FORMAT
ws2.cell(row=r_det_tot, column=9, value=f"=SUM(I5:I{r_det_tot-1})").number_format = CURRENCY_FORMAT
ws2.cell(row=r_det_tot, column=10, value=df_all['SALDO_AKHIR_MUTASI'].iloc[-1]).number_format = CURRENCY_FORMAT
ws2.cell(row=r_det_tot, column=11, value="-").alignment = align_center
ws2.cell(row=r_det_tot, column=12, value="-").alignment = align_center

for c in range(1, 13):
    cell = ws2.cell(row=r_det_tot, column=c)
    cell.font = font_bold
    cell.fill = fill_total
    cell.border = border_total
    if c in [8, 9, 10]:
        cell.alignment = align_right

# Freeze Panes for Detail Transaksi
ws2.freeze_panes = 'A5'

# ==============================================================================
# SHEET 3: REKAP BULANAN
# ==============================================================================
ws3 = wb.create_sheet(title='Rekap Bulanan')
ws3.views.sheetView[0].showGridLines = True

ws3['A1'] = "REKAPITULASI DETAIL PER BULAN"
ws3['A1'].font = font_title
ws3['A2'] = "Breakdown Transaksi & Transaksi Terbesar Per Bulan"
ws3['A2'].font = font_subtitle

headers_m_detail = ['Bulan', 'Saldo Awal', 'Total Uang Masuk', 'Total Uang Keluar', 'Net Cash Flow', 'Saldo Akhir', 'Jml Tran', 'Transaksi Masuk Terbesar', 'Nominal (Rp)', 'Transaksi Keluar Terbesar', 'Nominal (Rp)']
for col_num, h_text in enumerate(headers_m_detail, 1):
    cell = ws3.cell(row=4, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

for idx, f in enumerate(files, 1):
    m = f.split('_')[0]
    df_m = df_all[df_all['BULAN'] == m]
    r = idx + 4
    
    s_awal = df_m['SALDO_AWAL_MUTASI'].iloc[0]
    s_akhir = df_m['SALDO_AKHIR_MUTASI'].iloc[-1]
    
    # Top credit
    df_m_cr = df_m[df_m['MUTASI_KREDIT'] > 0]
    if len(df_m_cr) > 0:
        top_cr_row = df_m_cr.loc[df_m_cr['MUTASI_KREDIT'].idxmax()]
        top_cr_desc = top_cr_row['REMARK_CUSTOM'] if pd.notna(top_cr_row['REMARK_CUSTOM']) else top_cr_row['DESK_TRAN']
        top_cr_val = top_cr_row['MUTASI_KREDIT']
    else:
        top_cr_desc, top_cr_val = '-', 0.0

    # Top debit
    df_m_db = df_m[df_m['MUTASI_DEBET'] > 0]
    if len(df_m_db) > 0:
        top_db_row = df_m_db.loc[df_m_db['MUTASI_DEBET'].idxmax()]
        top_db_desc = top_db_row['REMARK_CUSTOM'] if pd.notna(top_db_row['REMARK_CUSTOM']) else top_db_row['DESK_TRAN']
        top_db_val = top_db_row['MUTASI_DEBET']
    else:
        top_db_desc, top_db_val = '-', 0.0

    month_names = {'2026-01':'Januari 2026', '2026-02':'Februari 2026', '2026-03':'Maret 2026', '2026-04':'April 2026', '2026-05':'Mei 2026', '2026-06':'Juni 2026', '2026-07':'Juli 2026'}
    
    ws3.cell(row=r, column=1, value=month_names.get(m, m)).alignment = align_left
    ws3.cell(row=r, column=2, value=s_awal).number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=3, value=f"=SUMIFS('Detail Transaksi'!I:I, 'Detail Transaksi'!B:B, \"{m}\")").number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=4, value=f"=SUMIFS('Detail Transaksi'!H:H, 'Detail Transaksi'!B:B, \"{m}\")").number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=6, value=s_akhir).number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=7, value=len(df_m)).number_format = NUMBER_FORMAT
    ws3.cell(row=r, column=8, value=str(top_cr_desc)[:45]).alignment = align_left
    ws3.cell(row=r, column=9, value=top_cr_val).number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=10, value=str(top_db_desc)[:45]).alignment = align_left
    ws3.cell(row=r, column=11, value=top_db_val).number_format = CURRENCY_FORMAT

    fill_cur = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
    for c in range(1, 12):
        cell = ws3.cell(row=r, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_cur.fill_type:
            cell.fill = fill_cur
        if c in [2, 3, 4, 5, 6, 9, 11]:
            cell.alignment = align_right
        elif c == 7:
            cell.alignment = align_center

# ==============================================================================
# SHEET 4: PAYMENT AKR
# ==============================================================================
ws4 = wb.create_sheet(title='Payment Akr')
ws4.views.sheetView[0].showGridLines = True

ws4['A1'] = "DETAIL TRANSFER MASUK - PT AKR CORPORINDO TBK"
ws4['A1'].font = font_title
ws4['A2'] = "No. Rekening: 13001002256309 | Periode: Januari - Juli 2026 | Rekap Penerimaan Penjualan AKR"
ws4['A2'].font = font_subtitle

# Filter dataframe for AKR incoming transfers
df_akr = df_all[(df_all['MUTASI_KREDIT'] > 0) & (
    df_all['DESK_TRAN'].astype(str).str.contains('AKR', case=False) | 
    df_all['REMARK_CUSTOM'].astype(str).str.contains('AKR', case=False)
)].copy()

# KPI Cards
# Card 1: Total Penerimaan AKR
ws4.merge_cells('A4:B4')
ws4['A4'] = "TOTAL PENERIMAAN AKR"
ws4['A4'].font = font_card_lbl
ws4['A4'].alignment = align_center
ws4['A4'].fill = fill_card
ws4.merge_cells('A5:B5')
r_akr_end = len(df_akr) + 7
ws4['A5'] = f"=SUM(H8:H{r_akr_end})"
ws4['A5'].font = Font(name=FONT_NAME, size=14, bold=True, color='2E75B6')
ws4['A5'].number_format = CURRENCY_FORMAT
ws4['A5'].alignment = align_center
ws4['A5'].fill = fill_card

# Card 2: Jumlah Transaksi
ws4.merge_cells('C4:D4')
ws4['C4'] = "JUMLAH TRANSAKSI"
ws4['C4'].font = font_card_lbl
ws4['C4'].alignment = align_center
ws4['C4'].fill = fill_card
ws4.merge_cells('C5:D5')
ws4['C5'] = len(df_akr)
ws4['C5'].font = font_card_num
ws4['C5'].number_format = NUMBER_FORMAT
ws4['C5'].alignment = align_center
ws4['C5'].fill = fill_card

# Card 3: Rata-Rata per Transaksi
ws4.merge_cells('E4:F4')
ws4['E4'] = "RATA-RATA PER TRANSAKSI"
ws4['E4'].font = font_card_lbl
ws4['E4'].alignment = align_center
ws4['E4'].fill = fill_card
ws4.merge_cells('E5:F5')
ws4['E5'] = f"=AVERAGE(H8:H{r_akr_end})"
ws4['E5'].font = font_card_num
ws4['E5'].number_format = CURRENCY_FORMAT
ws4['E5'].alignment = align_center
ws4['E5'].fill = fill_card

# Card 4: Transaksi Terbesar
ws4.merge_cells('G4:H4')
ws4['G4'] = "TRANSAKSI TERBESAR"
ws4['G4'].font = font_card_lbl
ws4['G4'].alignment = align_center
ws4['G4'].fill = fill_card
ws4.merge_cells('G5:H5')
ws4['G5'] = f"=MAX(H8:H{r_akr_end})"
ws4['G5'].font = font_card_num
ws4['G5'].number_format = CURRENCY_FORMAT
ws4['G5'].alignment = align_center
ws4['G5'].fill = fill_card

for row in range(4, 6):
    for col in range(1, 9):
        cell = ws4.cell(row=row, column=col)
        cell.border = border_card

# Table Header
headers_akr = [
    'No', 'Periode', 'Tanggal & Waktu', 'No. Rekening', 'Kategori',
    'Keterangan / Remark', 'Jenis (GL)', 'Mutasi Kredit / Masuk (Rp)',
    'User ID', 'Kode Tran', 'No. Invoice (Sheet1)'
]

inv_mapping_list = [
    "INV 464, 467",
    "INV 466, 469",
    "INV 470",
    "INV 465, 468, 471, 477",
    "INV 478, 480",
    "INV 004, 006",
    "INV 479, 482, 488, 489, 005, 007",
    "INV 481",
    "INV 008, 010, 011",
    "INV 012, 013, 015",
    "INV 014, 016, 018",
    "INV 017, 019",
    "INV 021, 022, 023",
    "INV 024, 025, 026",
    "INV 029",
    "INV 026",
    "INV 034, 035, 036, 037, 038, 039",
    "INV 040, 041",
    "INV 042",
    "INV 050, 051, 052, 053, 054",
    "INV 057, 058, 059, 060",
    "INV 064",
    "INV 070, 074, 075, 076, 077, 079, 081, 082, 083",
    "INV 096, 097",
    "INV 102, 103, 104, 105, 106, 107, 108",
    "INV 112, 114, 115, 117, 118, 119",
    "INV 125, 126, 127",
    "INV 129, 132, 135, 136, 137, 140, 141",
    "INV 133, 134",
    "INV 142",
    "INV 146",
    "INV 148, 149"
]

for col_num, h_text in enumerate(headers_akr, 1):
    cell = ws4.cell(row=7, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

# Populate Rows
row_idx = 8
for idx, (_, r) in enumerate(df_akr.iterrows(), 1):
    r_num = row_idx + idx - 1
    ws4.cell(row=r_num, column=1, value=idx).alignment = align_center
    ws4.cell(row=r_num, column=2, value=str(r['BULAN'])).alignment = align_center
    ws4.cell(row=r_num, column=3, value=str(r['TGL_TRAN'])).alignment = align_center
    ws4.cell(row=r_num, column=4, value=str(r['NOREK'])).alignment = align_center
    ws4.cell(row=r_num, column=5, value=str(r['KATEGORI'])).alignment = align_left
    
    rem_text = str(r['REMARK_CUSTOM']) if pd.notna(r['REMARK_CUSTOM']) and str(r['REMARK_CUSTOM']).strip() != "" else str(r['DESK_TRAN'])
    ws4.cell(row=r_num, column=6, value=rem_text).alignment = align_left
    ws4.cell(row=r_num, column=7, value='Kredit').alignment = align_center
    ws4.cell(row=r_num, column=8, value=r['MUTASI_KREDIT']).number_format = CURRENCY_FORMAT
    ws4.cell(row=r_num, column=9, value=str(r['TRUSER']) if pd.notna(r['TRUSER']) else '-').alignment = align_center
    ws4.cell(row=r_num, column=10, value=str(r['KODE_TRAN']) if pd.notna(r['KODE_TRAN']) else '-').alignment = align_center
    inv_str = inv_mapping_list[idx-1] if idx-1 < len(inv_mapping_list) else '-'
    ws4.cell(row=r_num, column=11, value=inv_str).alignment = align_left

    fill_row = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
    if r['MUTASI_KREDIT'] >= 100000000:
        fill_row = fill_green_tint

    for c in range(1, 12):
        cell = ws4.cell(row=r_num, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_row.fill_type:
            cell.fill = fill_row
        if c == 8:
            cell.alignment = align_right
            cell.font = font_kredit

# Total Row for AKR Sheet
r_akr_tot = row_idx + len(df_akr)
ws4.cell(row=r_akr_tot, column=1, value="").alignment = align_center
ws4.cell(row=r_akr_tot, column=2, value="TOTAL AKR").alignment = align_center
for c in range(3, 8):
    ws4.cell(row=r_akr_tot, column=c, value="-").alignment = align_center

ws4.cell(row=r_akr_tot, column=8, value=f"=SUM(H8:H{r_akr_tot-1})").number_format = CURRENCY_FORMAT
ws4.cell(row=r_akr_tot, column=9, value="-").alignment = align_center
ws4.cell(row=r_akr_tot, column=10, value="-").alignment = align_center
ws4.cell(row=r_akr_tot, column=11, value="-").alignment = align_center

for c in range(1, 12):
    cell = ws4.cell(row=r_akr_tot, column=c)
    cell.font = font_bold
    cell.fill = fill_total
    cell.border = border_total
    if c == 8:
        cell.alignment = align_right

ws4.freeze_panes = 'A8'

# ==============================================================================
# AUTO FIT COLUMN WIDTHS FOR ALL SHEETS
# ==============================================================================
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip title row and merged cells for length calc
            if cell.row in [1, 2, 4, 5] and sheet.title in ['Ringkasan Eksekutif', 'Payment Akr']:
                continue
            val_str = str(cell.value or '')
            if cell.number_format and 'Rp' in cell.number_format:
                val_str += ' Rp 999,999,999.00'
            max_len = max(max_len, len(val_str))
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

# Adjust specific column widths for polished layout
ws1.column_dimensions['A'].width = 38
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 26
ws1.column_dimensions['D'].width = 26
ws1.column_dimensions['E'].width = 24
ws1.column_dimensions['F'].width = 24
ws1.column_dimensions['G'].width = 18
ws1.column_dimensions['H'].width = 20

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 42
ws2.column_dimensions['F'].width = 50
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 24
ws2.column_dimensions['I'].width = 24
ws2.column_dimensions['J'].width = 24

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 20
ws3.column_dimensions['F'].width = 20
ws3.column_dimensions['G'].width = 12
ws3.column_dimensions['H'].width = 45
ws3.column_dimensions['I'].width = 22
ws3.column_dimensions['J'].width = 45
ws3.column_dimensions['K'].width = 22

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 42
ws4.column_dimensions['F'].width = 50
ws4.column_dimensions['G'].width = 12
ws4.column_dimensions['H'].width = 26
ws4.column_dimensions['I'].width = 15
ws4.column_dimensions['J'].width = 15

output_filename = "Rekap_Rekening_Koran_BRI_2026.xlsx"
wb.save(output_filename)
print(f"Successfully generated {output_filename}")
