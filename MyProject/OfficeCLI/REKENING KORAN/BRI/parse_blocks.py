import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('Akr_Payment.xlsx', data_only=True)
ws = wb['Sheet1']

# Let's inspect all tables in Sheet1.
# Tables appear across column ranges:
# Range 1: cols 1..7
# Range 2: cols 9..15
# Range 3: cols 17..23
# Range 4: cols 25..31

col_ranges = [(1, 7), (9, 15), (17, 23), (25, 31)]

blocks = []

# Scan for header rows where col_range[0] value is 'No' and col_range[0]+1 is 'Tanggal'
for r in range(1, ws.max_row + 1):
    for c_start, c_end in col_ranges:
        v_no = ws.cell(row=r, column=c_start).value
        v_tgl = ws.cell(row=r, column=c_start+1).value
        if str(v_no).strip() == 'No' and str(v_tgl).strip() == 'Tanggal':
            # Found a table header! Let's parse this table down until Bayar tgl / TOTAL
            invoices = []
            total_val = None
            bayar_tgl = None
            
            # Read rows below header
            curr_r = r + 1
            while curr_r <= ws.max_row + 5:
                # Check cells in this column block
                c_no = ws.cell(row=curr_r, column=c_start).value
                c_tgl = ws.cell(row=curr_r, column=c_start+1).value
                c_ket = ws.cell(row=curr_r, column=c_start+2).value
                c_jml = ws.cell(row=curr_r, column=c_start+6).value # col 7 relative
                
                # Check for invoice (when c_no is integer or valid row and c_tgl is present)
                if c_tgl is not None and str(c_tgl).strip() != '' and str(c_tgl).strip() != 'None':
                    tgl_str = str(c_tgl).strip()
                    # If it looks like an invoice no or number/code
                    if tgl_str not in ['No', 'Tanggal', 'TOTAL', 'Keterangan']:
                        invoices.append(tgl_str)
                
                # Check for TOTAL
                # Check across the relative columns
                for cc in range(c_start, c_end + 1):
                    val_cc = ws.cell(row=curr_r, column=cc).value
                    val_str = str(val_cc).strip()
                    if 'TOTAL' in val_str:
                        # total value might be in adjacent cell
                        for cc2 in range(cc, c_end + 1):
                            v2 = ws.cell(row=curr_r, column=cc2).value
                            if isinstance(v2, (int, float)):
                                total_val = v2
                    if 'Bayar tgl' in val_str:
                        bayar_tgl = val_str
                
                # If we encounter next 'No' header in same column block or reach empty block, break
                if curr_r > r + 1 and str(ws.cell(row=curr_r, column=c_start).value).strip() == 'No':
                    break
                # If we have passed 'Bayar tgl' or TOTAL and hit empty rows
                if bayar_tgl is not None and curr_r > r + 15:
                    break
                curr_r += 1

            blocks.append({
                'header_row': r,
                'col_start': c_start,
                'invoices': invoices,
                'total': total_val,
                'bayar_tgl': bayar_tgl
            })

print(f"Found {len(blocks)} blocks:")
for idx, b in enumerate(blocks, 1):
    print(f"Block {idx:2d} (R{b['header_row']}, C{b['col_start']}): Bayar={b['bayar_tgl']} | Total={b['total']} | Invoices={b['invoices']}")
