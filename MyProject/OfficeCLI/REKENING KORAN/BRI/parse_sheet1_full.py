import openpyxl
import re

wb = openpyxl.load_workbook('Akr_Payment.xlsx', data_only=True)
ws_sheet1 = wb['Sheet1']
ws_akr = wb['Payment Akr']

# Let's inspect Sheet1 structure by analyzing rows
# Each table block has headers like ('No', 'Tanggal', 'Keterangan', 'Jumlah')
# Let's find all header cells containing 'Tanggal'

tables = []

for r in range(1, ws_sheet1.max_row + 1):
    for c in range(1, 35):
        val = str(ws_sheet1.cell(row=r, column=c).value).strip()
        if val == 'Tanggal':
            # c is the column of 'Tanggal'.
            # c-1 is 'No' column
            # c+1 / c+2 is 'Keterangan' column
            # c+5 / c+6 is 'Jumlah' column
            
            # Let's find all rows for this table below row r
            inv_list = []
            total_amt = None
            bayar_tgl = None
            
            for rr in range(r + 1, r + 25):
                no_v = ws_sheet1.cell(row=rr, column=c-1).value
                tgl_v = ws_sheet1.cell(row=rr, column=c).value
                
                # Check for invoice in Tanggal column
                if tgl_v is not None and str(tgl_v).strip() != '' and str(tgl_v).strip() != 'None':
                    tstr = str(tgl_v).strip()
                    # Check if it's a date or invoice format or string
                    if tstr not in ['Tanggal', 'No', 'Keterangan', 'Jumlah', 'TOTAL']:
                        inv_list.append(tstr)
                
                # Check all cells in this block area for TOTAL and Bayar tgl
                for cc in range(max(1, c-2), c+8):
                    v_cell = str(ws_sheet1.cell(row=rr, column=cc).value or '').strip()
                    if 'TOTAL' in v_cell:
                        # find numerical value in this row
                        for cc2 in range(cc, cc+6):
                            v_num = ws_sheet1.cell(row=rr, column=cc2).value
                            if isinstance(v_num, (int, float)) and v_num > 0:
                                total_amt = v_num
                    if 'Bayar tgl' in v_cell:
                        bayar_tgl = v_cell
                        
            tables.append({
                'r': r,
                'c': c,
                'invoices': inv_list,
                'total': total_amt,
                'bayar_tgl': bayar_tgl
            })

with open('parsed_tables.txt', 'w') as f:
    f.write(f"Total tables found in Sheet1: {len(tables)}\n\n")
    for idx, t in enumerate(tables, 1):
        f.write(f"Table {idx:2d} [Row {t['r']}, Col {t['c']}]:\n")
        f.write(f"  Bayar tgl : {t['bayar_tgl']}\n")
        f.write(f"  Total     : {t['total']}\n")
        f.write(f"  Invoices  : {t['invoices']}\n\n")

print("Done parsing tables.")
