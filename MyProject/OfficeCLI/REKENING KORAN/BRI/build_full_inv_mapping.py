import pandas as pd
import openpyxl
import re

# Load Sheet1 and Payment Akr
wb = openpyxl.load_workbook('Akr_Payment.xlsx', data_only=True)
ws_s1 = wb['Sheet1']
ws_akr = wb['Payment Akr']

# Read df_akr
rows = []
for r in range(8, 40):
    no = ws_akr.cell(row=r, column=1).value
    bulan = ws_akr.cell(row=r, column=2).value
    tgl = ws_akr.cell(row=r, column=3).value
    norek = ws_akr.cell(row=r, column=4).value
    kat = ws_akr.cell(row=r, column=5).value
    rem = ws_akr.cell(row=r, column=6).value
    gl = ws_akr.cell(row=r, column=7).value
    amt = ws_akr.cell(row=r, column=8).value
    truser = ws_akr.cell(row=r, column=9).value
    kode = ws_akr.cell(row=r, column=10).value
    inv_old = ws_akr.cell(row=r, column=11).value
    rows.append({
        'r': r, 'no': no, 'bulan': bulan, 'tgl': tgl, 'norek': norek,
        'kat': kat, 'rem': rem, 'gl': gl, 'amt': amt, 'truser': truser,
        'kode': kode, 'inv_old': inv_old
    })

df_p = pd.DataFrame(rows)

# Parse Sheet1 blocks carefully
# A block in Sheet1 has:
# - Header row with 'No' and 'Tanggal'
# - List of rows with invoice numbers under 'Tanggal' column and descriptions/amounts
# - A 'TOTAL' row
# - A 'Bayar tgl :' row

s1_data = pd.read_excel('Akr_Payment.xlsx', sheet_name='Sheet1', header=None)

blocks = []
for r in range(len(s1_data)):
    for c in [0, 8, 16, 24]:
        if str(s1_data.iloc[r, c]).strip() == 'No' and str(s1_data.iloc[r, c+1]).strip() == 'Tanggal':
            invs = []
            details = []
            total_val = None
            bayar_tgl = None
            
            for rr in range(r + 1, min(r + 25, len(s1_data))):
                row_vals = s1_data.iloc[rr, c:c+8].tolist()
                tgl_val = str(row_vals[1]).strip() if pd.notna(row_vals[1]) else ''
                ket_val = str(row_vals[2]).strip() if pd.notna(row_vals[2]) else ''
                
                # Check for invoice in "Tanggal" column
                if tgl_val and tgl_val not in ['No', 'Tanggal', 'Keterangan', 'Jumlah', 'TOTAL', 'nan'] and not tgl_val.startswith('Bayar'):
                    invs.append(tgl_val)
                    if ket_val:
                        details.append(f"{tgl_val} ({ket_val})")
                
                for cc in range(c, min(c+8, s1_data.shape[1])):
                    cv = str(s1_data.iloc[rr, cc]).strip()
                    if 'TOTAL' in cv:
                        for cc2 in range(cc, min(cc+6, s1_data.shape[1])):
                            vnum = s1_data.iloc[rr, cc2]
                            if isinstance(vnum, (int, float)) and not pd.isna(vnum) and vnum > 0:
                                total_val = float(vnum)
                    if 'Bayar tgl' in cv:
                        bayar_tgl = cv
            
            blocks.append({
                'r': r, 'c': c, 'invs': invs, 'details': details, 'total': total_val, 'bayar_tgl': bayar_tgl
            })

print("=== PARSED BLOCKS FROM SHEET1 ===")
for i, b in enumerate(blocks, 1):
    print(f"B{i:2d} | Tgl: {b['bayar_tgl']} | Total: {b['total']} | Invs: {b['invs']}")
