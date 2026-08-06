import pandas as pd
import openpyxl

# Load data
wb_akr = openpyxl.load_workbook('Akr_Payment.xlsx', data_only=True)
ws_s1 = wb_akr['Sheet1']
ws_p = wb_akr['Payment Akr']

# Invoice mapping dictionary manually verified & matched with Sheet1 and bank remarks
# Key: Row index (0 to 31 in df_akr) or row 8 to 39 in ws_p
# Value: Detailed Invoice string

inv_mapping = {
    8:  "INV 464, 467",                          # 2026-01-05 (97,242,247) - Bank Permata remark
    9:  "INV 466, 469",                          # 2026-01-05 (117,154,200) - Bank Permata remark
    10: "INV 470",                               # 2026-01-08 (43,461,000) - Bank Permata remark
    11: "INV 465, 468, 471, 477",                # 2026-01-12 (109,603,000) - Sheet1 Table 1 (Bayar 12 Jan 2026)
    12: "INV 478, 480",                          # 2026-01-29 (53,153,200) - Bank Permata remark
    13: "INV 004, 006",                          # 2026-02-09 (21,589,166) - Bank Permata remark
    14: "INV 479, 482, 488, 489, 005, 007",      # 2026-02-09 (201,570,700) - Sheet1 Table 2 (Bayar 09 Feb 2026)
    15: "INV 481",                               # 2026-02-12 (28,457,200) - Bank Permata remark
    16: "INV 008, 010, 011",                     # 2026-02-19 (100,089,000) - Sheet1 Table 3 (Bayar 19 Feb 2026)
    17: "INV 012, 013, 015",                     # 2026-02-19 (174,810,400) - Bank Permata remark
    18: "INV 014, 016, 018",                     # 2026-03-02 (116,833,600) - Bank Permata remark
    19: "INV 017, 019",                          # 2026-03-02 (131,661,000) - Sheet1 Table 4 (Bayar 02 Mar 2026)
    20: "INV 021, 022, 023",                     # 2026-03-12 (98,294,278) - Bank Permata remark
    21: "INV 024, 025, 026",                     # 2026-03-16 (95,407,500) - Sheet1 Table 5 (Bayar 16 Mar 2026)
    22: "INV 029",                               # 2026-03-30 (42,442,181) - Bank Permata remark
    23: "INV 026",                               # 2026-03-30 (87,158,400) - Sheet1 Table 6 (Bayar 30 Mar 2026)
    24: "INV 034, 035, 036, 037, 038, 039",      # 2026-04-06 (103,584,000) - Bank Permata remark
    25: "INV 040, 041",                          # 2026-04-16 (31,887,200) - Bank Permata remark
    26: "INV 042",                               # 2026-04-27 (14,894,000) - Bank Permata remark
    27: "INV 050, 051, 052, 053, 054",           # 2026-05-18 (4,627,108) - Sheet1 Table 7 (Bayar 18 Mei 2026)
    28: "INV 057, 058, 059, 060",                # 2026-05-25 (45,184,500) - Sheet1 Table 8 (Bayar 25 Mei 2026)
    29: "INV 064",                               # 2026-05-29 (32,730,000) - Bank Permata remark
    30: "INV 070, 074, 075, 076, 077, 079, 081, 082, 083", # 2026-06-15 (213,928,000) - Sheet1 Table 9 & 10 (Bayar 15 Jun 2026)
    31: "INV 096, 097",                          # 2026-06-25 (32,674,089) - Bank Permata remark
    32: "INV 102, 103, 104, 105, 106, 107, 108", # 2026-07-02 (120,038,200) - Bank Permata remark
    33: "INV 112, 114, 115, 117, 118, 119",      # 2026-07-09 (100,377,375) - Bank Permata remark
    34: "INV 125, 126, 127",                     # 2026-07-16 (63,110,000) - Bank Permata remark
    35: "INV 129, 132, 135, 136, 137, 140, 141", # 2026-07-20 (145,563,000) - Sheet1 Table 11 (Bayar 20 Jul 2026)
    36: "INV 133, 134",                          # 2026-07-20 (13,326,000) - Bank Permata remark
    37: "INV 142",                               # 2026-07-23 (10,778,000) - Bank Permata remark
    38: "INV 146",                               # 2026-07-27 (2,203,000) - Bank Permata remark
    39: "INV 148, 149"                           # 2026-07-30 (33,866,800) - Bank Permata remark
}

print(f"Total rows mapped: {len(inv_mapping)}")

# Open workbook for editing (preserving formulas & structure)
wb_edit = openpyxl.load_workbook('Akr_Payment.xlsx')
ws_p_edit = wb_edit['Payment Akr']

# Check/Set Header for column 11 if not present
if ws_p_edit.cell(row=7, column=11).value is None or str(ws_p_edit.cell(row=7, column=11).value).strip() in ['', 'INV']:
    ws_p_edit.cell(row=7, column=11, value='No. Invoice / Remark (Sheet1)')

for r, inv_text in inv_mapping.items():
    ws_p_edit.cell(row=r, column=11, value=inv_text)
    # Also update Keterangan / Remark if needed, or keep column 11 clear
    print(f"Row {r:2d} -> {inv_text}")

wb_edit.save('Akr_Payment.xlsx')
print("Successfully updated Akr_Payment.xlsx!")
