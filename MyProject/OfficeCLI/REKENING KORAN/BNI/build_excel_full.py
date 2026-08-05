import glob, pdfplumber, re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def clean_page_chars(page):
    seen = set()
    clean = []
    for c in page.chars:
        key = (c['text'], round(c['x0'], 1), round(c['top'], 1))
        if key not in seen:
            seen.add(key)
            clean.append(c)
    return clean

def group_chars_by_line(chars):
    lines = []
    if not chars: return lines
    chars = sorted(chars, key=lambda c: (c['top'], c['x0']))
    cur = [chars[0]]
    for c in chars[1:]:
        if abs(c['top'] - cur[0]['top']) < 2.5:
            cur.append(c)
        else:
            lines.append(sorted(cur, key=lambda x: x['x0']))
            cur = [c]
    if cur:
        lines.append(sorted(cur, key=lambda x: x['x0']))
    return lines

def parse_pdf(filename):
    transactions = []
    meta = {}
    
    with pdfplumber.open(filename) as pdf:
        for page_idx, page in enumerate(pdf.pages, 1):
            chars = clean_page_chars(page)
            
            if page_idx == 1:
                page_text = page.extract_text()
                for line in page_text.split('\n'):
                    if 'Ledger Balance:' in line:
                        m = re.search(r'Ledger Balance:\s*([\d,]+\.\d{2})', line)
                        if m: meta['ledger_balance'] = float(m.group(1).replace(',', ''))
                    if 'Period' in line:
                        meta['period'] = line.split(':', 1)[-1].strip() if ':' in line else line
            
            last_text = page.extract_text()
            for line in last_text.split('\n'):
                if 'Ending Balance' in line:
                    m = re.search(r'Ending Balance\s*:\s*([\d,]+\.\d{2})', line)
                    if m: meta['ending_balance'] = float(m.group(1).replace(',', ''))
                if 'Total Debet' in line:
                    m = re.search(r'Total Debet\s*:\s*(\d+)?\s*([\d,]+\.\d{2})', line)
                    if m: meta['total_debit'] = float(m.group(2).replace(',', ''))
                if 'Total Credit' in line:
                    m = re.search(r'Total Credit\s*:\s*(\d+)?\s*([\d,]+\.\d{2})', line)
                    if m: meta['total_credit'] = float(m.group(2).replace(',', ''))

            lines_dict = {}
            for c in chars:
                if c['top'] < 320: continue
                t = round(c['top'], 1)
                matched = None
                for k in lines_dict:
                    if abs(k - t) < 2.5:
                        matched = k
                        break
                if matched is None:
                    lines_dict[t] = [c]
                else:
                    lines_dict[matched].append(c)
            
            sorted_tops = sorted(lines_dict.keys())
            table_tops = []
            for t in sorted_tops:
                l_str = ''.join(c['text'] for c in lines_dict[t]).strip()
                if any(k in l_str for k in ['Ending Balance', 'Total Debet', 'Total Credit', 'Ledger Balance']):
                    continue
                if 'SEJAHTERA BERSAMA 0' in l_str or 'ACCOUNT STATEMENT' in l_str:
                    continue
                table_tops.append(t)
                
            anchors = []
            for t in table_tops:
                l_chars = sorted(lines_dict[t], key=lambda x: x['x0'])
                d_str = ''.join(c['text'] for c in l_chars if 15 <= c['x0'] <= 125).strip()
                if re.match(r'^\d{2}/\d{2}/\d{4}', d_str):
                    anchors.append(t)
                    
            if not anchors:
                continue
                
            tx_groups = {anc: [] for anc in anchors}
            for t in table_tops:
                closest_anc = min(anchors, key=lambda a: abs(a - t))
                tx_groups[closest_anc].append(t)
                
            for anc in anchors:
                tops_for_tx = sorted(tx_groups[anc])
                tx_chars = [c for t in tops_for_tx for c in lines_dict[t]]
                
                anc_chars = sorted(lines_dict[anc], key=lambda c: c['x0'])
                post_date = ''.join(c['text'] for c in anc_chars if 15 <= c['x0'] < 125).strip()
                eff_date = ''.join(c['text'] for c in anc_chars if 125 <= c['x0'] < 230).strip()
                journal = ''.join(c['text'] for c in anc_chars if 290 <= c['x0'] < 335).strip()
                db_cr = ''.join(c['text'] for c in anc_chars if 575 <= c['x0'] < 610).strip()
                balance_str = ''.join(c['text'] for c in anc_chars if 610 <= c['x0'] < 710).strip()
                
                if not balance_str:
                    bal_chars = sorted([c for c in tx_chars if 610 <= c['x0'] < 710], key=lambda c: (c['top'], c['x0']))
                    balance_str = ''.join(c['text'] for c in bal_chars).strip()
                
                branch_chars = [c for c in tx_chars if 230 <= c['x0'] < 290]
                branch_lines = group_chars_by_line(branch_chars)
                branch = ' '.join(''.join(c['text'] for c in l) for l in branch_lines).strip()
                
                desc_chars = [c for c in tx_chars if 335 <= c['x0'] < 500]
                desc_lines = group_chars_by_line(desc_chars)
                desc = ' '.join(''.join(c['text'] for c in l) for l in desc_lines).strip()
                
                balance_val = float(balance_str.replace(',', '')) if balance_str else None
                
                transactions.append({
                    'source_file': filename,
                    'page': page_idx,
                    'posting_date': post_date,
                    'effective_date': eff_date,
                    'branch': branch,
                    'journal': journal,
                    'description': desc,
                    'db_cr': db_cr,
                    'balance_raw': balance_str,
                    'balance': balance_val
                })
                
    return meta, transactions

def get_category(desc):
    d = desc.upper()
    if 'MPN G2' in d: return 'Pajak (MPN G2)'
    if 'BPJS KES' in d: return 'BPJS Kesehatan'
    if 'BPJS TK' in d: return 'BPJS Ketenagakerjaan'
    if 'PLN' in d or 'BIAYA ADMIN (PLN' in d: return 'Listrik / PLN'
    if 'BIFAST' in d or 'BI FAST' in d: return 'Transfer BI-FAST'
    if 'SETOR TUNAI' in d: return 'Setor Tunai'
    if 'JASA GIRO' in d: return 'Jasa Giro'
    if 'BIAYA ADM' in d or 'BY TRX ATM' in d: return 'Biaya Admin / Bank'
    if 'PPH' in d: return 'Pajak PPh'
    if 'PEMINDAHAN' in d or 'TRANSFER' in d: return 'Transfer / Pemindahan'
    return 'Lainnya'

month_names = [
    ('2026-01_RK_BNI.pdf', 'Jan 2026'),
    ('2026-02_RK_BNI.pdf', 'Feb 2026'),
    ('2026-03_RK_BNI.pdf', 'Mar 2026'),
    ('2026-04_RK_BNI.pdf', 'Apr 2026'),
    ('2026-05_RK_BNI.pdf', 'Mei 2026'),
    ('2026-06_RK_BNI.pdf', 'Jun 2026'),
    ('2026-07_RK_BNI.pdf', 'Jul 2026'),
]

monthly_data = []
all_transactions = []

global_tx_counter = 1

for f, m_name in month_names:
    meta, txs = parse_pdf(f)
    cur_bal = meta['ledger_balance']
    for tx in txs:
        if tx['balance'] is not None:
            diff = round(tx['balance'] - cur_bal, 2)
            tx['amount'] = abs(diff)
            tx['db_cr'] = 'K' if diff > 0 else 'D'
            cur_bal = tx['balance']
        else:
            diff = round(meta['ending_balance'] - cur_bal, 2)
            tx['amount'] = abs(diff)
            tx['db_cr'] = 'K' if diff > 0 else 'D'
            tx['balance'] = meta['ending_balance']
            cur_bal = meta['ending_balance']
        tx['month'] = m_name
        tx['category'] = get_category(tx['description'])
        tx['global_id'] = global_tx_counter
        global_tx_counter += 1
        all_transactions.append(tx)
        
    monthly_data.append({
        'filename': f,
        'month': m_name,
        'meta': meta,
        'txs': txs,
        'ledger_balance': meta['ledger_balance'],
        'ending_balance': meta['ending_balance'],
        'total_debit': meta['total_debit'],
        'total_credit': meta['total_credit'],
        'count': len(txs)
    })

# Build Workbook
wb = openpyxl.Workbook()
wb.remove(wb.active) # remove default sheet

# Fonts & Styles
font_name = "Calibri"

title_font = Font(name=font_name, size=16, bold=True, color="005E6A")
subtitle_font = Font(name=font_name, size=11, italic=True, color="555555")
section_font = Font(name=font_name, size=12, bold=True, color="004B54")
header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
kpi_title_font = Font(name=font_name, size=9, bold=True, color="555555")
kpi_val_font = Font(name=font_name, size=14, bold=True, color="004B54")
kpi_val_green = Font(name=font_name, size=14, bold=True, color="1E8449")
kpi_val_red = Font(name=font_name, size=14, bold=True, color="C0392B")

data_font = Font(name=font_name, size=10)
bold_data_font = Font(name=font_name, size=10, bold=True)
italic_font = Font(name=font_name, size=9, italic=True, color="777777")

header_fill = PatternFill(start_color="005E6A", end_color="005E6A", fill_type="solid")
header_fill_sec = PatternFill(start_color="00838F", end_color="00838F", fill_type="solid")
sub_header_fill = PatternFill(start_color="E6F2F4", end_color="E6F2F4", fill_type="solid")
zebra_fill = PatternFill(start_color="F9FCFC", end_color="F9FCFC", fill_type="solid")
total_fill = PatternFill(start_color="E0F0F2", end_color="E0F0F2", fill_type="solid")

kpi_fill_init = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
kpi_fill_in = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
kpi_fill_out = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
kpi_fill_end = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
kpi_fill_net = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

thin_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
thick_top = Side(border_style="thin", color="005E6A")
double_bottom = Side(border_style="double", color="005E6A")
total_border = Border(top=thick_top, bottom=double_bottom, left=thin_side, right=thin_side)
kpi_border = Border(left=Side(border_style="thin", color="B0BEC5"), right=Side(border_style="thin", color="B0BEC5"),
                    top=Side(border_style="thin", color="B0BEC5"), bottom=Side(border_style="thin", color="B0BEC5"))

curr_fmt = '"Rp "#,##0.00'
int_fmt = '#,##0'

# ==========================================
# 1. SHEET: DASHBOARD
# ==========================================
ws_dash = wb.create_sheet(title="Dashboard Ringkasan")
ws_dash.views.sheetView[0].showGridLines = True

# Title block
ws_dash.merge_cells("A1:G1")
ws_dash["A1"] = "REKAPITULASI TRANSAKSI REKENING KORAN BNI"
ws_dash["A1"].font = title_font
ws_dash["A1"].alignment = Alignment(vertical="center")

ws_dash.merge_cells("A2:G2")
ws_dash["A2"] = "PT HIJAU SEJAHTERA BERSAMA | Rekening Giro No: 692291557 | Periode: 01 Jan 2026 - 31 Jul 2026"
ws_dash["A2"].font = subtitle_font

# Metadata Box
ws_dash["A4"] = "Informasi Perusahaan & Rekening"
ws_dash["A4"].font = section_font

meta_labels = [
    ("Nama Perusahaan", "PT HIJAU SEJAHTERA BERSAMA"),
    ("Nomor Rekening", "692291557"),
    ("Mata Uang / Tipe", "IDR / CURRENT (Giro BNI)"),
    ("Alamat Perusahaan", "DUSUN I PUJOSARI RT 001 RW 001"),
    ("Periode Transaksi", "01 Januari 2026 - 31 Juli 2026 (7 Bulan)"),
]

for idx, (label, val) in enumerate(meta_labels, start=5):
    ws_dash[f"A{idx}"] = label
    ws_dash[f"A{idx}"].font = bold_data_font
    ws_dash[f"A{idx}"].fill = sub_header_fill
    ws_dash[f"A{idx}"].border = thin_border
    
    ws_dash[f"B{idx}"] = val
    ws_dash[f"B{idx}"].font = data_font
    ws_dash[f"B{idx}"].border = thin_border
    ws_dash.merge_cells(f"B{idx}:D{idx}")

# KPI Cards Block (Columns F to K or D to H)
# Card 1: Saldo Awal (01 Jan 2026)
ws_dash["F4"] = "SALDO AWAL (01 JAN 2026)"
ws_dash["F4"].font = kpi_title_font
ws_dash["F4"].fill = kpi_fill_init
ws_dash["F4"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["F4"].border = kpi_border

ws_dash["F5"] = monthly_data[0]['ledger_balance']
ws_dash["F5"].font = kpi_val_font
ws_dash["F5"].fill = kpi_fill_init
ws_dash["F5"].number_format = curr_fmt
ws_dash["F5"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["F5"].border = kpi_border

# Card 2: Total Kredit (Pemasukan)
ws_dash["G4"] = "TOTAL PEMASUKAN (KREDIT)"
ws_dash["G4"].font = kpi_title_font
ws_dash["G4"].fill = kpi_fill_in
ws_dash["G4"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["G4"].border = kpi_border

ws_dash["G5"] = "=SUM(E12:E18)"
ws_dash["G5"].font = kpi_val_green
ws_dash["G5"].fill = kpi_fill_in
ws_dash["G5"].number_format = curr_fmt
ws_dash["G5"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["G5"].border = kpi_border

# Card 3: Total Debet (Pengeluaran)
ws_dash["H4"] = "TOTAL PENGELUARAN (DEBET)"
ws_dash["H4"].font = kpi_title_font
ws_dash["H4"].fill = kpi_fill_out
ws_dash["H4"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["H4"].border = kpi_border

ws_dash["H5"] = "=SUM(D12:D18)"
ws_dash["H5"].font = kpi_val_red
ws_dash["H5"].fill = kpi_fill_out
ws_dash["H5"].number_format = curr_fmt
ws_dash["H5"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["H5"].border = kpi_border

# Card 4: Net Cash Flow
ws_dash["I4"] = "NET CASH FLOW"
ws_dash["I4"].font = kpi_title_font
ws_dash["I4"].fill = kpi_fill_net
ws_dash["I4"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["I4"].border = kpi_border

ws_dash["I5"] = "=G5-H5"
ws_dash["I5"].font = kpi_val_font
ws_dash["I5"].fill = kpi_fill_net
ws_dash["I5"].number_format = curr_fmt
ws_dash["I5"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["I5"].border = kpi_border

# Card 5: Saldo Akhir (31 Jul 2026)
ws_dash["J4"] = "SALDO AKHIR (31 JUL 2026)"
ws_dash["J4"].font = kpi_title_font
ws_dash["J4"].fill = kpi_fill_end
ws_dash["J4"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["J4"].border = kpi_border

ws_dash["J5"] = monthly_data[-1]['ending_balance']
ws_dash["J5"].font = kpi_val_font
ws_dash["J5"].fill = kpi_fill_end
ws_dash["J5"].number_format = curr_fmt
ws_dash["J5"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["J5"].border = kpi_border

# Monthly Summary Table
ws_dash["A10"] = "Ringkasan Mutasi Per Bulan (Jan - Jul 2026)"
ws_dash["A10"].font = section_font

headers_summary = [
    "No", "Bulan", "Saldo Awal (Rp)", "Total Debet / Keluar (Rp)", "Total Kredit / Masuk (Rp)", 
    "Net Flow (Rp)", "Saldo Akhir (Rp)", "Jumlah Transaksi"
]

for col_num, h_text in enumerate(headers_summary, start=1):
    cell = ws_dash.cell(row=11, column=col_num, value=h_text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

ws_dash.row_dimensions[11].height = 28

for idx, m_info in enumerate(monthly_data, start=12):
    r_idx = idx
    ws_dash.cell(row=r_idx, column=1, value=idx-11).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=r_idx, column=2, value=m_info['month']).alignment = Alignment(horizontal="center")
    
    # Saldo Awal
    c_start = ws_dash.cell(row=r_idx, column=3, value=m_info['ledger_balance'])
    c_start.number_format = curr_fmt
    
    # Total Debet
    c_deb = ws_dash.cell(row=r_idx, column=4, value=m_info['total_debit'])
    c_deb.number_format = curr_fmt
    
    # Total Kredit
    c_cred = ws_dash.cell(row=r_idx, column=5, value=m_info['total_credit'])
    c_cred.number_format = curr_fmt
    
    # Net Flow
    c_net = ws_dash.cell(row=r_idx, column=6, value=f"=E{r_idx}-D{r_idx}")
    c_net.number_format = curr_fmt
    
    # Saldo Akhir
    c_end = ws_dash.cell(row=r_idx, column=7, value=m_info['ending_balance'])
    c_end.number_format = curr_fmt
    
    # Jml Transaksi
    c_cnt = ws_dash.cell(row=r_idx, column=8, value=m_info['count'])
    c_cnt.number_format = int_fmt
    c_cnt.alignment = Alignment(horizontal="center")
    
    fill_row = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    for col_c in range(1, 9):
        c_item = ws_dash.cell(row=r_idx, column=col_c)
        c_item.font = data_font
        if fill_row.fill_type: c_item.fill = fill_row
        c_item.border = thin_border

# Totals Row for Summary
tot_row = 19
ws_dash.cell(row=tot_row, column=1, value="")
ws_dash.cell(row=tot_row, column=2, value="TOTAL / SALDO").alignment = Alignment(horizontal="center")
ws_dash.cell(row=tot_row, column=3, value=monthly_data[0]['ledger_balance']).number_format = curr_fmt
ws_dash.cell(row=tot_row, column=4, value="=SUM(D12:D18)").number_format = curr_fmt
ws_dash.cell(row=tot_row, column=5, value="=SUM(E12:E18)").number_format = curr_fmt
ws_dash.cell(row=tot_row, column=6, value="=E19-D19").number_format = curr_fmt
ws_dash.cell(row=tot_row, column=7, value=monthly_data[-1]['ending_balance']).number_format = curr_fmt
ws_dash.cell(row=tot_row, column=8, value="=SUM(H12:H18)").number_format = int_fmt
ws_dash.cell(row=tot_row, column=8).alignment = Alignment(horizontal="center")

for col_c in range(1, 9):
    c_item = ws_dash.cell(row=tot_row, column=col_c)
    c_item.font = bold_data_font
    c_item.fill = total_fill
    c_item.border = total_border

# Category Breakdown Table
ws_dash["A22"] = "Breakdown Transaksi Berdasarkan Kategori Pengeluaran & Pemasukan"
ws_dash["A22"].font = section_font

headers_cat = ["No", "Kategori Transaksi", "Jumlah Transaksi", "Total Debet / Pengeluaran (Rp)", "Total Kredit / Pemasukan (Rp)", "Net Impact (Rp)"]

for col_num, h_text in enumerate(headers_cat, start=1):
    cell = ws_dash.cell(row=23, column=col_num, value=h_text)
    cell.font = header_font
    cell.fill = header_fill_sec
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

ws_dash.row_dimensions[23].height = 26

categories = [
    'Transfer / Pemindahan',
    'Pajak (MPN G2)',
    'BPJS Kesehatan',
    'BPJS Ketenagakerjaan',
    'Listrik / PLN',
    'Transfer BI-FAST',
    'Setor Tunai',
    'Jasa Giro',
    'Biaya Admin / Bank',
    'Pajak PPh',
    'Lainnya'
]

cat_start_row = 24
for idx, cat_name in enumerate(categories, start=cat_start_row):
    r_idx = idx
    ws_dash.cell(row=r_idx, column=1, value=idx-cat_start_row+1).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=r_idx, column=2, value=cat_name)
    
    # Formulas linking to 'Rekap Master' sheet
    c_cnt = ws_dash.cell(row=r_idx, column=3, value=f"=COUNTIF('Rekap Master'!H:H, B{r_idx})")
    c_cnt.number_format = int_fmt
    c_cnt.alignment = Alignment(horizontal="center")
    
    c_deb = ws_dash.cell(row=r_idx, column=4, value=f"=SUMIF('Rekap Master'!H:H, B{r_idx}, 'Rekap Master'!J:J)")
    c_deb.number_format = curr_fmt
    
    c_cred = ws_dash.cell(row=r_idx, column=5, value=f"=SUMIF('Rekap Master'!H:H, B{r_idx}, 'Rekap Master'!K:K)")
    c_cred.number_format = curr_fmt
    
    c_net = ws_dash.cell(row=r_idx, column=6, value=f"=E{r_idx}-D{r_idx}")
    c_net.number_format = curr_fmt
    
    fill_row = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    for col_c in range(1, 7):
        c_item = ws_dash.cell(row=r_idx, column=col_c)
        c_item.font = data_font
        if fill_row.fill_type: c_item.fill = fill_row
        c_item.border = thin_border

cat_tot_row = cat_start_row + len(categories)
ws_dash.cell(row=cat_tot_row, column=1, value="")
ws_dash.cell(row=cat_tot_row, column=2, value="TOTAL KATEGORI").alignment = Alignment(horizontal="center")
ws_dash.cell(row=cat_tot_row, column=3, value=f"=SUM(C{cat_start_row}:C{cat_tot_row-1})").number_format = int_fmt
ws_dash.cell(row=cat_tot_row, column=3).alignment = Alignment(horizontal="center")
ws_dash.cell(row=cat_tot_row, column=4, value=f"=SUM(D{cat_start_row}:D{cat_tot_row-1})").number_format = curr_fmt
ws_dash.cell(row=cat_tot_row, column=5, value=f"=SUM(E{cat_start_row}:E{cat_tot_row-1})").number_format = curr_fmt
ws_dash.cell(row=cat_tot_row, column=6, value=f"=E{cat_tot_row}-D{cat_tot_row}").number_format = curr_fmt

for col_c in range(1, 7):
    c_item = ws_dash.cell(row=cat_tot_row, column=col_c)
    c_item.font = bold_data_font
    c_item.fill = total_fill
    c_item.border = total_border


# ==========================================
# 2. SHEET: REKAP MASTER (ALL TRANSACTIONS)
# ==========================================
ws_master = wb.create_sheet(title="Rekap Master")
ws_master.views.sheetView[0].showGridLines = True

ws_master.merge_cells("A1:L1")
ws_master["A1"] = "DAFTAR MASTER TRANSAKSI REKENING KORAN BNI (JANUARI - JULI 2026)"
ws_master["A1"].font = title_font

ws_master.merge_cells("A2:L2")
ws_master["A2"] = "PT HIJAU SEJAHTERA BERSAMA | Total Transaksi: 129 Records"
ws_master["A2"].font = subtitle_font

headers_master = [
    "No", "Bulan", "Posting Date", "Effective Date", "No. Jurnal", "Cabang / Branch", 
    "Uraian / Deskripsi Transaksi", "Kategori", "DB/CR", "Debet / Keluar (Rp)", "Kredit / Masuk (Rp)", "Saldo Akhir (Rp)"
]

for col_num, h_text in enumerate(headers_master, start=1):
    cell = ws_master.cell(row=4, column=col_num, value=h_text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

ws_master.row_dimensions[4].height = 28

for idx, tx in enumerate(all_transactions, start=5):
    r_idx = idx
    ws_master.cell(row=r_idx, column=1, value=tx['global_id']).alignment = Alignment(horizontal="center")
    ws_master.cell(row=r_idx, column=2, value=tx['month']).alignment = Alignment(horizontal="center")
    ws_master.cell(row=r_idx, column=3, value=tx['posting_date']).alignment = Alignment(horizontal="center")
    ws_master.cell(row=r_idx, column=4, value=tx['effective_date']).alignment = Alignment(horizontal="center")
    ws_master.cell(row=r_idx, column=5, value=tx['journal']).alignment = Alignment(horizontal="center")
    ws_master.cell(row=r_idx, column=6, value=tx['branch'])
    ws_master.cell(row=r_idx, column=7, value=tx['description'])
    ws_master.cell(row=r_idx, column=8, value=tx['category'])
    
    db_cr_cell = ws_master.cell(row=r_idx, column=9, value=tx['db_cr'])
    db_cr_cell.alignment = Alignment(horizontal="center")
    if tx['db_cr'] == 'D':
        db_cr_cell.font = Font(name=font_name, size=10, bold=True, color="C0392B")
    else:
        db_cr_cell.font = Font(name=font_name, size=10, bold=True, color="1E8449")
        
    deb_val = tx['amount'] if tx['db_cr'] == 'D' else 0.0
    cred_val = tx['amount'] if tx['db_cr'] == 'K' else 0.0
    
    c_deb = ws_master.cell(row=r_idx, column=10, value=deb_val)
    c_deb.number_format = curr_fmt
    
    c_cred = ws_master.cell(row=r_idx, column=11, value=cred_val)
    c_cred.number_format = curr_fmt
    
    c_bal = ws_master.cell(row=r_idx, column=12, value=tx['balance'])
    c_bal.number_format = curr_fmt
    
    fill_row = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    for col_c in range(1, 13):
        c_item = ws_master.cell(row=r_idx, column=col_c)
        if col_c != 9: c_item.font = data_font
        if fill_row.fill_type: c_item.fill = fill_row
        c_item.border = thin_border

# Total row Master
m_tot_row = len(all_transactions) + 5
ws_master.cell(row=m_tot_row, column=1, value="")
ws_master.cell(row=m_tot_row, column=2, value="TOTAL").alignment = Alignment(horizontal="center")
ws_master.cell(row=m_tot_row, column=10, value=f"=SUM(J5:J{m_tot_row-1})").number_format = curr_fmt
ws_master.cell(row=m_tot_row, column=11, value=f"=SUM(K5:K{m_tot_row-1})").number_format = curr_fmt
ws_master.cell(row=m_tot_row, column=12, value=all_transactions[-1]['balance']).number_format = curr_fmt

for col_c in range(1, 13):
    c_item = ws_master.cell(row=m_tot_row, column=col_c)
    c_item.font = bold_data_font
    c_item.fill = total_fill
    c_item.border = total_border

ws_master.freeze_panes = "A5"
ws_master.auto_filter.ref = f"A4:L{m_tot_row-1}"


# ==========================================
# 3. SHEETS: MONTHLY DETAIL TABS
# ==========================================
for m_info in monthly_data:
    m_name = m_info['month']
    meta = m_info['meta']
    txs = m_info['txs']
    f_name = m_info['filename']
    
    ws_m = wb.create_sheet(title=m_name)
    ws_m.views.sheetView[0].showGridLines = True
    
    ws_m.merge_cells("A1:K1")
    ws_m["A1"] = f"REKAP TRANSAKSI REKENING KORAN BNI - PERIODE {m_name.upper()}"
    ws_m["A1"].font = title_font
    
    ws_m.merge_cells("A2:K2")
    ws_m["A2"] = f"PT HIJAU SEJAHTERA BERSAMA | Periode: {meta.get('period', m_name)} | Source: {f_name}"
    ws_m["A2"].font = subtitle_font
    
    # Monthly Header Stats Cards
    ws_m["A4"] = "SALDO AWAL"
    ws_m["A4"].font = kpi_title_font; ws_m["A4"].fill = kpi_fill_init; ws_m["A4"].alignment = Alignment(horizontal="center"); ws_m["A4"].border = kpi_border
    ws_m["A5"] = meta['ledger_balance']
    ws_m["A5"].font = kpi_val_font; ws_m["A5"].fill = kpi_fill_init; ws_m["A5"].number_format = curr_fmt; ws_m["A5"].alignment = Alignment(horizontal="center"); ws_m["A5"].border = kpi_border
    
    ws_m["C4"] = "TOTAL DEBET (KELUAR)"
    ws_m["C4"].font = kpi_title_font; ws_m["C4"].fill = kpi_fill_out; ws_m["C4"].alignment = Alignment(horizontal="center"); ws_m["C4"].border = kpi_border
    ws_m["C5"] = f"=SUM(I8:I{len(txs)+7})"
    ws_m["C5"].font = kpi_val_red; ws_m["C5"].fill = kpi_fill_out; ws_m["C5"].number_format = curr_fmt; ws_m["C5"].alignment = Alignment(horizontal="center"); ws_m["C5"].border = kpi_border
    
    ws_m["E4"] = "TOTAL KREDIT (MASUK)"
    ws_m["E4"].font = kpi_title_font; ws_m["E4"].fill = kpi_fill_in; ws_m["E4"].alignment = Alignment(horizontal="center"); ws_m["E4"].border = kpi_border
    ws_m["E5"] = f"=SUM(J8:J{len(txs)+7})"
    ws_m["E5"].font = kpi_val_green; ws_m["E5"].fill = kpi_fill_in; ws_m["E5"].number_format = curr_fmt; ws_m["E5"].alignment = Alignment(horizontal="center"); ws_m["E5"].border = kpi_border
    
    ws_m["G4"] = "SALDO AKHIR"
    ws_m["G4"].font = kpi_title_font; ws_m["G4"].fill = kpi_fill_end; ws_m["G4"].alignment = Alignment(horizontal="center"); ws_m["G4"].border = kpi_border
    ws_m["G5"] = meta['ending_balance']
    ws_m["G5"].font = kpi_val_font; ws_m["G5"].fill = kpi_fill_end; ws_m["G5"].number_format = curr_fmt; ws_m["G5"].alignment = Alignment(horizontal="center"); ws_m["G5"].border = kpi_border
    
    headers_detail = [
        "No", "Posting Date", "Effective Date", "No. Jurnal", "Cabang / Branch", 
        "Uraian / Deskripsi Transaksi", "Kategori", "DB/CR", "Debet / Keluar (Rp)", "Kredit / Masuk (Rp)", "Saldo Akhir (Rp)"
    ]
    
    for col_num, h_text in enumerate(headers_detail, start=1):
        cell = ws_m.cell(row=7, column=col_num, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws_m.row_dimensions[7].height = 28
    
    for idx, tx in enumerate(txs, start=8):
        r_idx = idx
        ws_m.cell(row=r_idx, column=1, value=idx-7).alignment = Alignment(horizontal="center")
        ws_m.cell(row=r_idx, column=2, value=tx['posting_date']).alignment = Alignment(horizontal="center")
        ws_m.cell(row=r_idx, column=3, value=tx['effective_date']).alignment = Alignment(horizontal="center")
        ws_m.cell(row=r_idx, column=4, value=tx['journal']).alignment = Alignment(horizontal="center")
        ws_m.cell(row=r_idx, column=5, value=tx['branch'])
        ws_m.cell(row=r_idx, column=6, value=tx['description'])
        ws_m.cell(row=r_idx, column=7, value=tx['category'])
        
        db_cr_cell = ws_m.cell(row=r_idx, column=8, value=tx['db_cr'])
        db_cr_cell.alignment = Alignment(horizontal="center")
        if tx['db_cr'] == 'D':
            db_cr_cell.font = Font(name=font_name, size=10, bold=True, color="C0392B")
        else:
            db_cr_cell.font = Font(name=font_name, size=10, bold=True, color="1E8449")
            
        deb_val = tx['amount'] if tx['db_cr'] == 'D' else 0.0
        cred_val = tx['amount'] if tx['db_cr'] == 'K' else 0.0
        
        c_deb = ws_m.cell(row=r_idx, column=9, value=deb_val)
        c_deb.number_format = curr_fmt
        
        c_cred = ws_m.cell(row=r_idx, column=10, value=cred_val)
        c_cred.number_format = curr_fmt
        
        c_bal = ws_m.cell(row=r_idx, column=11, value=tx['balance'])
        c_bal.number_format = curr_fmt
        
        fill_row = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)
        for col_c in range(1, 12):
            c_item = ws_m.cell(row=r_idx, column=col_c)
            if col_c != 8: c_item.font = data_font
            if fill_row.fill_type: c_item.fill = fill_row
            c_item.border = thin_border
            
    mon_tot_row = len(txs) + 8
    ws_m.cell(row=mon_tot_row, column=1, value="")
    ws_m.cell(row=mon_tot_row, column=2, value="TOTAL").alignment = Alignment(horizontal="center")
    ws_m.cell(row=mon_tot_row, column=9, value=f"=SUM(I8:I{mon_tot_row-1})").number_format = curr_fmt
    ws_m.cell(row=mon_tot_row, column=10, value=f"=SUM(J8:J{mon_tot_row-1})").number_format = curr_fmt
    ws_m.cell(row=mon_tot_row, column=11, value=meta['ending_balance']).number_format = curr_fmt
    
    for col_c in range(1, 12):
        c_item = ws_m.cell(row=mon_tot_row, column=col_c)
        c_item.font = bold_data_font
        c_item.fill = total_fill
        c_item.border = total_border
        
    ws_m.freeze_panes = "A8"
    ws_m.auto_filter.ref = f"A7:K{mon_tot_row-1}"


# Auto-fit Column Widths across all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Skip checking merged title row or long headers
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row in [1, 2, 4, 5] and sheet.title == "Dashboard Ringkasan":
                continue
            if cell.row in [1, 2] and sheet.title != "Dashboard Ringkasan":
                continue
            if len(val_str) > max_len and len(val_str) < 80:
                max_len = len(val_str)
                
        adjusted_width = max(max_len + 4, 12)
        sheet.column_dimensions[col_letter].width = min(adjusted_width, 60)

output_filename = "Rekap_Transaksi_BNI_2026.xlsx"
wb.save(output_filename)
print(f"Workbook successfully saved to: {output_filename}")
