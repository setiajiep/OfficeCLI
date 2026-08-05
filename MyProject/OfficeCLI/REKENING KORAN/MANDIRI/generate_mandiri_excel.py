import os, glob
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Path & Load Data
path = '/root/MyProject/OfficeCLI/REKENING KORAN/MANDIRI'
files = [
    ('account_statement_1140016063946_01 January 2026-31 January 2026_20260805140110.xlsx', '2026-01', 'Januari 2026'),
    ('account_statement_1140016063946_01 February 2026-28 February 2026_20260805140325.xlsx', '2026-02', 'Februari 2026'),
    ('account_statement_1140016063946_01 March 2026-31 March 2026_20260805140420.xlsx', '2026-03', 'Maret 2026'),
    ('account_statement_1140016063946_01 April 2026-30 April 2026_20260805140524.xlsx', '2026-04', 'April 2026'),
    ('account_statement_1140016063946_01 May 2026-31 May 2026_20260805140642.xlsx', '2026-05', 'Mei 2026'),
    ('account_statement_1140016063946_01 June 2026-30 June 2026_20260805140816.xlsx', '2026-06', 'Juni 2026'),
    ('account_statement_1140016063946_01 July 2026-31 July 2026_20260805140941.xlsx', '2026-07', 'Juli 2026')
]

dfs = []
for fname, code_m, name_m in files:
    fpath = os.path.join(path, fname)
    df = pd.read_excel(fpath)
    df_valid = df[df['Date'].notna()].copy()
    df_valid['PERIODE_CODE'] = code_m
    df_valid['PERIODE_NAME'] = name_m
    dfs.append(df_valid)

df_all = pd.concat(dfs, ignore_index=True)
df_all['Debit'] = df_all['Debit'].astype(float).fillna(0.0)
df_all['Credit'] = df_all['Credit'].astype(float).fillna(0.0)

def categorize_mandiri(row):
    d = (str(row['Description']) + ' ' + str(row['Description.1'])).upper()
    deb = float(row['Debit'])
    kre = float(row['Credit'])
    
    if kre > 0:
        if 'BUNGA' in d:
            return 'Pendapatan Bunga Bank'
        elif 'MANAMBANG MUARA ENIM' in d:
            return 'Penerimaan Penjualan (PT Manambang Muara Enim)'
        elif 'AKR CORPORINDO' in d:
            return 'Penerimaan Penjualan (PT AKR Corporindo)'
        elif 'BIO NUSANTARA' in d:
            return 'Penerimaan Penjualan (PT Bio Nusantara)'
        elif 'SRI KARYA LINTASINDO' in d:
            return 'Penerimaan Penjualan (PT Sri Karya Lintasindo)'
        elif 'NOAHTU SHIPYARD' in d:
            return 'Penerimaan Penjualan / Refund (PT Noahtu Shipyard)'
        elif 'DUA PUTRA PERKASA' in d:
            return 'Penerimaan Penjualan (PT Dua Putra Perkasa Pratama)'
        elif 'HILMAWAN INDRA' in d:
            return 'Transfer Masuk Pihak Ketiga (Hilmawan Indra Waskita)'
        elif 'BURHAN' in d:
            return 'Transfer Masuk Internal (Burhan)'
        elif 'HIJAU SEJAHTERA BERSAMA' in d or 'BRINIDJA/PT HIJAU' in d or 'BNINIDJA/HIJAU' in d or 'CENAIDJA/HIJAU' in d:
            return 'Pemindahbukuan Rekening Internal (HSB)'
        elif 'PRMA CR TRANSF' in d:
            return 'Transfer Masuk Pihak Ketiga (PRMA/ATM)'
        else:
            return 'Transfer Masuk Lainnya'
    else:
        if 'AUTO COLL' in d:
            return 'Pembayaran Angsuran Pinjaman (Auto Coll)'
        elif 'AKR CORPORINDO' in d:
            return 'Pembayaran Pembelian (PT AKR Corporindo)'
        elif 'BURHAN' in d:
            return 'Transfer Keluar Internal (Burhan)'
        elif 'WIJI INDRAYANI' in d:
            return 'Transfer Keluar (Wiji Indrayani)'
        elif 'TARIK TUNAI' in d:
            return 'Penarikan Tunai Cek'
        elif 'NOAHTU SHIPYARD' in d:
            return 'Pembayaran Vendor (PT Noahtu Shipyard)'
        elif 'LAUTAN BERLIAN' in d:
            return 'Pembayaran Vendor (PT Lautan Berlian)'
        elif 'HIJAU SEJAHTERA BERSAMA' in d or 'BRINIDJA/PT HIJAU' in d or 'BNINIDJA/HIJAU' in d or 'CENAIDJA/HIJAU' in d:
            return 'Pemindahbukuan Rekening Internal (HSB)'
        elif 'BIAYA ADM' in d or 'BUKU CEK' in d or 'METERAI' in d:
            return 'Biaya Administrasi Bank & Buku Cek'
        elif deb == 2500:
            return 'Biaya Admin Transfer'
        elif 'PAJAK' in d:
            return 'Pajak Bunga Tabungan'
        else:
            return 'Pengeluaran Lainnya'

df_all['KATEGORI'] = df_all.apply(categorize_mandiri, axis=1)

# Create Workbook
wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

FONT_NAME = 'Segoe UI'

font_title = Font(name=FONT_NAME, size=16, bold=True, color='0F4C81') # Mandiri Navy Blue
font_subtitle = Font(name=FONT_NAME, size=10, italic=True, color='595959')
font_section = Font(name=FONT_NAME, size=12, bold=True, color='0F4C81')
font_header = Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')
font_card_num = Font(name=FONT_NAME, size=14, bold=True, color='0F4C81')
font_card_lbl = Font(name=FONT_NAME, size=9, bold=True, color='595959')

font_bold = Font(name=FONT_NAME, size=10, bold=True)
font_regular = Font(name=FONT_NAME, size=10)
font_kredit = Font(name=FONT_NAME, size=10, color='1E7E34') # Rich green
font_debet = Font(name=FONT_NAME, size=10, color='BD2130') # Rich red

fill_navy = PatternFill(start_color='0F4C81', end_color='0F4C81', fill_type='solid') # Mandiri Blue
fill_card = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
fill_green_tint = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')
fill_red_tint = PatternFill(start_color='FCE8E6', end_color='FCE8E6', fill_type='solid')
fill_total = PatternFill(start_color='D0E1F9', end_color='D0E1F9', fill_type='solid')

thin_border = Side(border_style='thin', color='D1D5DB')
thick_bottom = Side(border_style='medium', color='0F4C81')
double_bottom = Side(border_style='double', color='0F4C81')
top_thin = Side(border_style='thin', color='0F4C81')

border_all_thin = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
border_header = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thick_bottom)
border_total = Border(top=top_thin, bottom=double_bottom, left=thin_border, right=thin_border)
border_card = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

CURRENCY_FORMAT = 'Rp #,##0.00;[Red](Rp #,##0.00);"-"'
NUMBER_FORMAT = '#,##0'

# ==============================================================================
# SHEET 1: RINGKASAN EKSEKUTIF
# ==============================================================================
ws1 = wb.create_sheet(title='Ringkasan Eksekutif')
ws1.views.sheetView[0].showGridLines = True

ws1['A1'] = "REKAPITULASI REKENING KORAN BANK MANDIRI"
ws1['A1'].font = font_title
ws1['A2'] = "No. Rekening: 1140016063946 | Pemilik: PT HIJAU SEJAHTERA BERSAMA | Periode: Januari - Juli 2026 | Mata Uang: IDR"
ws1['A2'].font = font_subtitle

# KPI Cards
# Card 1: Total Transaksi
ws1.merge_cells('A4:B4')
ws1['A4'] = "TOTAL TRANSAKSI"
ws1['A4'].font = font_card_lbl
ws1['A4'].alignment = align_center
ws1['A4'].fill = fill_card
ws1.merge_cells('A5:B5')
ws1['A5'] = len(df_all)
ws1['A5'].font = font_card_num
ws1['A5'].number_format = NUMBER_FORMAT
ws1['A5'].alignment = align_center
ws1['A5'].fill = fill_card

# Card 2: Total Uang Masuk (Kredit)
ws1.merge_cells('C4:D4')
ws1['C4'] = "TOTAL UANG MASUK (KREDIT)"
ws1['C4'].font = font_card_lbl
ws1['C4'].alignment = align_center
ws1['C4'].fill = fill_card
ws1.merge_cells('C5:D5')
ws1['C5'] = df_all['Credit'].sum()
ws1['C5'].font = Font(name=FONT_NAME, size=14, bold=True, color='1E7E34')
ws1['C5'].number_format = CURRENCY_FORMAT
ws1['C5'].alignment = align_center
ws1['C5'].fill = fill_card

# Card 3: Total Uang Keluar (Debet)
ws1.merge_cells('E4:F4')
ws1['E4'] = "TOTAL UANG KELUAR (DEBET)"
ws1['E4'].font = font_card_lbl
ws1['E4'].alignment = align_center
ws1['E4'].fill = fill_card
ws1.merge_cells('E5:F5')
ws1['E5'] = df_all['Debit'].sum()
ws1['E5'].font = Font(name=FONT_NAME, size=14, bold=True, color='BD2130')
ws1['E5'].number_format = CURRENCY_FORMAT
ws1['E5'].alignment = align_center
ws1['E5'].fill = fill_card

# Card 4: Net Cash Flow
ws1.merge_cells('G4:H4')
ws1['G4'] = "NET CASH FLOW (SURPLUS)"
ws1['G4'].font = font_card_lbl
ws1['G4'].alignment = align_center
ws1['G4'].fill = fill_card
ws1.merge_cells('G5:H5')
ws1['G5'] = "=C5-E5"
ws1['G5'].font = font_card_num
ws1['G5'].number_format = CURRENCY_FORMAT
ws1['G5'].alignment = align_center
ws1['G5'].fill = fill_card

for row in range(4, 6):
    for col in range(1, 9):
        cell = ws1.cell(row=row, column=col)
        cell.border = border_card

# Section 1: Tabel Ringkasan Bulanan
ws1['A7'] = "1. RINGKASAN ARUS KAS BULANAN (JANUARI - JULI 2026)"
ws1['A7'].font = font_section

headers_m = ['Periode Bulan', 'Total Masuk / Kredit (Rp)', 'Total Keluar / Debet (Rp)', 'Net Cash Flow (Rp)', 'Jumlah Transaksi', 'Rata-Rata Transaksi']
for col_num, h_text in enumerate(headers_m, 1):
    cell = ws1.cell(row=8, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

row_start = 9
for idx, (fname, code_m, name_m) in enumerate(files):
    r = row_start + idx
    ws1.cell(row=r, column=1, value=name_m).alignment = align_left
    ws1.cell(row=r, column=2, value=f"=SUMIFS('Detail Transaksi'!I:I, 'Detail Transaksi'!B:B, \"{code_m}\")").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=3, value=f"=SUMIFS('Detail Transaksi'!H:H, 'Detail Transaksi'!B:B, \"{code_m}\")").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=5, value=f"=COUNTIFS('Detail Transaksi'!B:B, \"{code_m}\")").number_format = NUMBER_FORMAT
    ws1.cell(row=r, column=6, value=f"=(B{r}+C{r})/E{r}").number_format = CURRENCY_FORMAT

    fill_cur = fill_zebra if idx % 2 == 1 else PatternFill(fill_type=None)
    for c in range(1, 7):
        cell = ws1.cell(row=r, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_cur.fill_type:
            cell.fill = fill_cur
        if c in [2, 3, 4, 6]:
            cell.alignment = align_right
        elif c == 5:
            cell.alignment = align_center

# Total Row for Monthly Summary
r_tot = row_start + len(files)
ws1.cell(row=r_tot, column=1, value="TOTAL ARUS KAS").alignment = align_left
ws1.cell(row=r_tot, column=2, value=f"=SUM(B9:B{r_tot-1})").number_format = CURRENCY_FORMAT
ws1.cell(row=r_tot, column=3, value=f"=SUM(C9:C{r_tot-1})").number_format = CURRENCY_FORMAT
ws1.cell(row=r_tot, column=4, value=f"=B{r_tot}-C{r_tot}").number_format = CURRENCY_FORMAT
ws1.cell(row=r_tot, column=5, value=f"=SUM(E9:E{r_tot-1})").number_format = NUMBER_FORMAT
ws1.cell(row=r_tot, column=6, value=f"=(B{r_tot}+C{r_tot})/E{r_tot}").number_format = CURRENCY_FORMAT

for c in range(1, 7):
    cell = ws1.cell(row=r_tot, column=c)
    cell.font = font_bold
    cell.fill = fill_total
    cell.border = border_total
    if c in [2, 3, 4, 6]:
        cell.alignment = align_right
    elif c == 5:
        cell.alignment = align_center

# Section 2: Tabel Rekapitulasi Kategori Transaksi
r_cat_head = r_tot + 3
ws1.cell(row=r_cat_head-1, column=1, value="2. REKAPITULASI KATEGORI TRANSAKSI").font = font_section

headers_cat = ['No', 'Kategori Transaksi', 'Jenis Arus Kas', 'Total Debet (Pengeluaran)', 'Total Kredit (Penerimaan)', 'Net Cash Impact', 'Jumlah Transaksi', '% Dari Total Mutasi']
for col_num, h_text in enumerate(headers_cat, 1):
    cell = ws1.cell(row=r_cat_head, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

categories_list = [
    # Uang Masuk
    ('Penerimaan Penjualan (PT Manambang Muara Enim)', 'Uang Masuk'),
    ('Penerimaan Penjualan (PT Bio Nusantara)', 'Uang Masuk'),
    ('Penerimaan Penjualan (PT Sri Karya Lintasindo)', 'Uang Masuk'),
    ('Penerimaan Penjualan / Refund (PT Noahtu Shipyard)', 'Uang Masuk'),
    ('Penerimaan Penjualan (PT Dua Putra Perkasa Pratama)', 'Uang Masuk'),
    ('Penerimaan Penjualan (PT AKR Corporindo)', 'Uang Masuk'),
    ('Transfer Masuk Internal (Burhan)', 'Uang Masuk'),
    ('Transfer Masuk Pihak Ketiga (Hilmawan Indra Waskita)', 'Uang Masuk'),
    ('Transfer Masuk Pihak Ketiga (PRMA/ATM)', 'Uang Masuk'),
    ('Pendapatan Bunga Bank', 'Uang Masuk'),
    # Uang Keluar
    ('Transfer Keluar Internal (Burhan)', 'Uang Keluar'),
    ('Penarikan Tunai Cek', 'Uang Keluar'),
    ('Pemindahbukuan Rekening Internal (HSB)', 'Net (Masuk/Keluar)'),
    ('Transfer Keluar (Wiji Indrayani)', 'Uang Keluar'),
    ('Pembayaran Pembelian (PT AKR Corporindo)', 'Uang Keluar'),
    ('Pembayaran Angsuran Pinjaman (Auto Coll)', 'Uang Keluar'),
    ('Pembayaran Vendor (PT Lautan Berlian)', 'Uang Keluar'),
    ('Biaya Administrasi Bank & Buku Cek', 'Uang Keluar'),
    ('Biaya Admin Transfer', 'Uang Keluar'),
    ('Pajak Bunga Tabungan', 'Uang Keluar')
]

r_cat_start = r_cat_head + 1
for idx, (cat_name, cat_type) in enumerate(categories_list, 1):
    r = r_cat_start + idx - 1
    ws1.cell(row=r, column=1, value=idx).alignment = align_center
    ws1.cell(row=r, column=2, value=cat_name).alignment = align_left
    ws1.cell(row=r, column=3, value=cat_type).alignment = align_center
    ws1.cell(row=r, column=4, value=f"=SUMIFS('Detail Transaksi'!H:H, 'Detail Transaksi'!E:E, B{r})").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=5, value=f"=SUMIFS('Detail Transaksi'!I:I, 'Detail Transaksi'!E:E, B{r})").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=6, value=f"=E{r}-D{r}").number_format = CURRENCY_FORMAT
    ws1.cell(row=r, column=7, value=f"=COUNTIFS('Detail Transaksi'!E:E, B{r})").number_format = NUMBER_FORMAT
    ws1.cell(row=r, column=8, value=f"=(D{r}+E{r})/('Ringkasan Eksekutif'!C5+'Ringkasan Eksekutif'!E5)").number_format = '0.00%'

    fill_cur = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
    for c in range(1, 9):
        cell = ws1.cell(row=r, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_cur.fill_type:
            cell.fill = fill_cur
        if c in [4, 5, 6, 8]:
            cell.alignment = align_right

# Category Total Row
r_cat_tot = r_cat_start + len(categories_list)
ws1.cell(row=r_cat_tot, column=1, value="").alignment = align_center
ws1.cell(row=r_cat_tot, column=2, value="TOTAL KATEGORI").alignment = align_left
ws1.cell(row=r_cat_tot, column=3, value="-").alignment = align_center
ws1.cell(row=r_cat_tot, column=4, value=f"=SUM(D{r_cat_start}:D{r_cat_tot-1})").number_format = CURRENCY_FORMAT
ws1.cell(row=r_cat_tot, column=5, value=f"=SUM(E{r_cat_start}:E{r_cat_tot-1})").number_format = CURRENCY_FORMAT
ws1.cell(row=r_cat_tot, column=6, value=f"=E{r_cat_tot}-D{r_cat_tot}").number_format = CURRENCY_FORMAT
ws1.cell(row=r_cat_tot, column=7, value=f"=SUM(G{r_cat_start}:G{r_cat_tot-1})").number_format = NUMBER_FORMAT
ws1.cell(row=r_cat_tot, column=8, value=f"=SUM(H{r_cat_start}:H{r_cat_tot-1})").number_format = '0.00%'

for c in range(1, 9):
    cell = ws1.cell(row=r_cat_tot, column=c)
    cell.font = font_bold
    cell.fill = fill_total
    cell.border = border_total
    if c in [4, 5, 6, 8]:
        cell.alignment = align_right
    elif c == 7:
        cell.alignment = align_center

# ==============================================================================
# SHEET 2: DETAIL TRANSAKSI
# ==============================================================================
ws2 = wb.create_sheet(title='Detail Transaksi')
ws2.views.sheetView[0].showGridLines = True

ws2['A1'] = "DETAIL MUTASI TRANSAKSI REKENING KORAN MANDIRI"
ws2['A1'].font = font_title
ws2['A2'] = "No. Rekening: 1140016063946 | Total Transaksi: 214 Items"
ws2['A2'].font = font_subtitle

headers_det = [
    'No', 'Periode', 'Tanggal Mutasi', 'No. Rekening', 'Kategori',
    'Deskripsi 1', 'Deskripsi 2 / Remark', 'Mutasi Debet (Rp)',
    'Mutasi Kredit (Rp)', 'Kode Tran', 'No. Referensi'
]

for col_num, h_text in enumerate(headers_det, 1):
    cell = ws2.cell(row=4, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

for idx, r in df_all.iterrows():
    row_num = idx + 5
    ws2.cell(row=row_num, column=1, value=idx+1).alignment = align_center
    ws2.cell(row=row_num, column=2, value=str(r['PERIODE_CODE'])).alignment = align_center
    ws2.cell(row=row_num, column=3, value=str(r['Date'])).alignment = align_center
    ws2.cell(row=row_num, column=4, value="1140016063946").alignment = align_center
    ws2.cell(row=row_num, column=5, value=str(r['KATEGORI'])).alignment = align_left
    
    ws2.cell(row=row_num, column=6, value=str(r['Description']) if pd.notna(r['Description']) else '').alignment = align_left
    ws2.cell(row=row_num, column=7, value=str(r['Description.1']) if pd.notna(r['Description.1']) else '').alignment = align_left
    
    ws2.cell(row=row_num, column=8, value=r['Debit']).number_format = CURRENCY_FORMAT
    ws2.cell(row=row_num, column=9, value=r['Credit']).number_format = CURRENCY_FORMAT
    
    ws2.cell(row=row_num, column=10, value=str(int(r['Transaction Code'])) if pd.notna(r['Transaction Code']) else '-').alignment = align_center
    ws2.cell(row=row_num, column=11, value=str(r['Reference No.']) if pd.notna(r['Reference No.']) else '-').alignment = align_center

    # Row styling & highlight
    fill_row = fill_zebra if idx % 2 == 1 else PatternFill(fill_type=None)
    if r['Credit'] >= 500000000:
        fill_row = fill_green_tint
    elif r['Debit'] >= 500000000:
        fill_row = fill_red_tint

    for c in range(1, 12):
        cell = ws2.cell(row=row_num, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_row.fill_type:
            cell.fill = fill_row
        if c in [8, 9]:
            cell.alignment = align_right
            if c == 9 and r['Credit'] > 0:
                cell.font = font_kredit
            elif c == 8 and r['Debit'] > 0:
                cell.font = font_debet

# Total Row for Detail
r_det_tot = len(df_all) + 5
ws2.cell(row=r_det_tot, column=1, value="").alignment = align_center
ws2.cell(row=r_det_tot, column=2, value="TOTAL").alignment = align_center
for c in range(3, 8):
    ws2.cell(row=r_det_tot, column=c, value="-").alignment = align_center

ws2.cell(row=r_det_tot, column=8, value=f"=SUM(H5:H{r_det_tot-1})").number_format = CURRENCY_FORMAT
ws2.cell(row=r_det_tot, column=9, value=f"=SUM(I5:I{r_det_tot-1})").number_format = CURRENCY_FORMAT
ws2.cell(row=r_det_tot, column=10, value="-").alignment = align_center
ws2.cell(row=r_det_tot, column=11, value="-").alignment = align_center

for c in range(1, 12):
    cell = ws2.cell(row=r_det_tot, column=c)
    cell.font = font_bold
    cell.fill = fill_total
    cell.border = border_total
    if c in [8, 9]:
        cell.alignment = align_right

ws2.freeze_panes = 'A5'

# ==============================================================================
# SHEET 3: REKAP BULANAN
# ==============================================================================
ws3 = wb.create_sheet(title='Rekap Bulanan')
ws3.views.sheetView[0].showGridLines = True

ws3['A1'] = "REKAPITULASI DETAIL PER BULAN"
ws3['A1'].font = font_title
ws3['A2'] = "Breakdown Mutasi & Transaksi Terbesar Per Bulan (Januari - Juli 2026)"
ws3['A2'].font = font_subtitle

headers_m_detail = ['Bulan', 'Total Uang Masuk', 'Total Uang Keluar', 'Net Cash Flow', 'Jml Tran', 'Transaksi Masuk Terbesar', 'Nominal (Rp)', 'Transaksi Keluar Terbesar', 'Nominal (Rp)']
for col_num, h_text in enumerate(headers_m_detail, 1):
    cell = ws3.cell(row=4, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

for idx, (fname, code_m, name_m) in enumerate(files, 1):
    df_m = df_all[df_all['PERIODE_CODE'] == code_m]
    r = idx + 4
    
    # Top credit
    df_m_cr = df_m[df_m['Credit'] > 0]
    if len(df_m_cr) > 0:
        top_cr_row = df_m_cr.loc[df_m_cr['Credit'].idxmax()]
        top_cr_desc = f"{top_cr_row['Description']} {top_cr_row['Description.1']}".strip()
        top_cr_val = top_cr_row['Credit']
    else:
        top_cr_desc, top_cr_val = '-', 0.0

    # Top debit
    df_m_db = df_m[df_m['Debit'] > 0]
    if len(df_m_db) > 0:
        top_db_row = df_m_db.loc[df_m_db['Debit'].idxmax()]
        top_db_desc = f"{top_db_row['Description']} {top_db_row['Description.1']}".strip()
        top_db_val = top_db_row['Debit']
    else:
        top_db_desc, top_db_val = '-', 0.0

    ws3.cell(row=r, column=1, value=name_m).alignment = align_left
    ws3.cell(row=r, column=2, value=f"=SUMIFS('Detail Transaksi'!I:I, 'Detail Transaksi'!B:B, \"{code_m}\")").number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=3, value=f"=SUMIFS('Detail Transaksi'!H:H, 'Detail Transaksi'!B:B, \"{code_m}\")").number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=5, value=len(df_m)).number_format = NUMBER_FORMAT
    ws3.cell(row=r, column=6, value=str(top_cr_desc)[:45]).alignment = align_left
    ws3.cell(row=r, column=7, value=top_cr_val).number_format = CURRENCY_FORMAT
    ws3.cell(row=r, column=8, value=str(top_db_desc)[:45]).alignment = align_left
    ws3.cell(row=r, column=9, value=top_db_val).number_format = CURRENCY_FORMAT

    fill_cur = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
    for c in range(1, 10):
        cell = ws3.cell(row=r, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_cur.fill_type:
            cell.fill = fill_cur
        if c in [2, 3, 4, 7, 9]:
            cell.alignment = align_right
        elif c == 5:
            cell.alignment = align_center

# ==============================================================================
# SHEET 4: PENERIMAAN PENJUALAN
# ==============================================================================
ws4 = wb.create_sheet(title='Penerimaan Penjualan')
ws4.views.sheetView[0].showGridLines = True

ws4['A1'] = "DETAIL PENERIMAAN PENJUALAN & CLIENT TRANSFERS"
ws4['A1'].font = font_title
ws4['A2'] = "No. Rekening: 1140016063946 | Periode: Januari - Juli 2026 | Rekapitulasi Pembayaran Pelanggan Utam"
ws4['A2'].font = font_subtitle

# Filter sales income (Manambang Muara Enim, Bio Nusantara, AKR, Sri Karya Lintasindo, Dua Putra, Noahtu)
sales_cats = [
    'Penerimaan Penjualan (PT Manambang Muara Enim)',
    'Penerimaan Penjualan (PT Bio Nusantara)',
    'Penerimaan Penjualan (PT Sri Karya Lintasindo)',
    'Penerimaan Penjualan / Refund (PT Noahtu Shipyard)',
    'Penerimaan Penjualan (PT Dua Putra Perkasa Pratama)',
    'Penerimaan Penjualan (PT AKR Corporindo)'
]
df_sales = df_all[df_all['KATEGORI'].isin(sales_cats)].copy()

# KPI Cards
# Card 1: Total Penerimaan Penjualan
ws4.merge_cells('A4:B4')
ws4['A4'] = "TOTAL PENERIMAAN PENJUALAN"
ws4['A4'].font = font_card_lbl
ws4['A4'].alignment = align_center
ws4['A4'].fill = fill_card
ws4.merge_cells('A5:B5')
r_sales_end = len(df_sales) + 7
ws4['A5'] = f"=SUM(H8:H{r_sales_end})"
ws4['A5'].font = Font(name=FONT_NAME, size=14, bold=True, color='1E7E34')
ws4['A5'].number_format = CURRENCY_FORMAT
ws4['A5'].alignment = align_center
ws4['A5'].fill = fill_card

# Card 2: Jumlah Transaksi Penjualan
ws4.merge_cells('C4:D4')
ws4['C4'] = "JUMLAH TRANSAKSI"
ws4['C4'].font = font_card_lbl
ws4['C4'].alignment = align_center
ws4['C4'].fill = fill_card
ws4.merge_cells('C5:D5')
ws4['C5'] = len(df_sales)
ws4['C5'].font = font_card_num
ws4['C5'].number_format = NUMBER_FORMAT
ws4['C5'].alignment = align_center
ws4['C5'].fill = fill_card

# Card 3: Transaksi Penjualan Terbesar
ws4.merge_cells('E4:F4')
ws4['E4'] = "TRANSAKSI TERBESAR"
ws4['E4'].font = font_card_lbl
ws4['E4'].alignment = align_center
ws4['E4'].fill = fill_card
ws4.merge_cells('E5:F5')
ws4['E5'] = f"=MAX(H8:H{r_sales_end})"
ws4['E5'].font = font_card_num
ws4['E5'].number_format = CURRENCY_FORMAT
ws4['E5'].alignment = align_center
ws4['E5'].fill = fill_card

# Card 4: Rata-Rata per Transaksi
ws4.merge_cells('G4:H4')
ws4['G4'] = "RATA-RATA PER TRANSAKSI"
ws4['G4'].font = font_card_lbl
ws4['G4'].alignment = align_center
ws4['G4'].fill = fill_card
ws4.merge_cells('G5:H5')
ws4['G5'] = f"=AVERAGE(H8:H{r_sales_end})"
ws4['G5'].font = font_card_num
ws4['G5'].number_format = CURRENCY_FORMAT
ws4['G5'].alignment = align_center
ws4['G5'].fill = fill_card

for row in range(4, 6):
    for col in range(1, 9):
        cell = ws4.cell(row=row, column=col)
        cell.border = border_card

# Table Header
headers_sales = [
    'No', 'Periode', 'Tanggal Mutasi', 'No. Rekening', 'Pelanggan / Kategori',
    'Deskripsi 1', 'Deskripsi 2 / Remark', 'Nominal Kredit / Masuk (Rp)',
    'Kode Tran', 'No. Referensi'
]

for col_num, h_text in enumerate(headers_sales, 1):
    cell = ws4.cell(row=7, column=col_num, value=h_text)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_header
    cell.border = border_header

# Populate Rows
row_idx = 8
for idx, (_, r) in enumerate(df_sales.iterrows(), 1):
    r_num = row_idx + idx - 1
    ws4.cell(row=r_num, column=1, value=idx).alignment = align_center
    ws4.cell(row=r_num, column=2, value=str(r['PERIODE_CODE'])).alignment = align_center
    ws4.cell(row=r_num, column=3, value=str(r['Date'])).alignment = align_center
    ws4.cell(row=r_num, column=4, value="1140016063946").alignment = align_center
    ws4.cell(row=r_num, column=5, value=str(r['KATEGORI'])).alignment = align_left
    
    ws4.cell(row=r_num, column=6, value=str(r['Description']) if pd.notna(r['Description']) else '').alignment = align_left
    ws4.cell(row=r_num, column=7, value=str(r['Description.1']) if pd.notna(r['Description.1']) else '').alignment = align_left
    ws4.cell(row=r_num, column=8, value=r['Credit']).number_format = CURRENCY_FORMAT
    
    ws4.cell(row=r_num, column=9, value=str(int(r['Transaction Code'])) if pd.notna(r['Transaction Code']) else '-').alignment = align_center
    ws4.cell(row=r_num, column=10, value=str(r['Reference No.']) if pd.notna(r['Reference No.']) else '-').alignment = align_center

    fill_row = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
    if r['Credit'] >= 1000000000:
        fill_row = fill_green_tint

    for c in range(1, 11):
        cell = ws4.cell(row=r_num, column=c)
        cell.font = font_regular
        cell.border = border_all_thin
        if fill_row.fill_type:
            cell.fill = fill_row
        if c == 8:
            cell.alignment = align_right
            cell.font = font_kredit

# Total Row for Sales Sheet
r_sales_tot = row_idx + len(df_sales)
ws4.cell(row=r_sales_tot, column=1, value="").alignment = align_center
ws4.cell(row=r_sales_tot, column=2, value="TOTAL PENERIMAAN").alignment = align_center
for c in range(3, 8):
    ws4.cell(row=r_sales_tot, column=c, value="-").alignment = align_center

ws4.cell(row=r_sales_tot, column=8, value=f"=SUM(H8:H{r_sales_tot-1})").number_format = CURRENCY_FORMAT
ws4.cell(row=r_sales_tot, column=9, value="-").alignment = align_center
ws4.cell(row=r_sales_tot, column=10, value="-").alignment = align_center

for c in range(1, 11):
    cell = ws4.cell(row=r_sales_tot, column=c)
    cell.font = font_bold
    cell.fill = fill_total
    cell.border = border_total
    if c == 8:
        cell.alignment = align_right

ws4.freeze_panes = 'A8'

# Column width tuning
ws1.column_dimensions['A'].width = 45
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 28
ws1.column_dimensions['D'].width = 25
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 22

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 48
ws2.column_dimensions['F'].width = 50
ws2.column_dimensions['G'].width = 45
ws2.column_dimensions['H'].width = 25
ws2.column_dimensions['I'].width = 25
ws2.column_dimensions['J'].width = 14
ws2.column_dimensions['K'].width = 28

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 45
ws3.column_dimensions['G'].width = 22
ws3.column_dimensions['H'].width = 45
ws3.column_dimensions['I'].width = 22

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 48
ws4.column_dimensions['F'].width = 50
ws4.column_dimensions['G'].width = 45
ws4.column_dimensions['H'].width = 26
ws4.column_dimensions['I'].width = 14
ws4.column_dimensions['J'].width = 28

output_filename = os.path.join(path, "Rekap_Rekening_Koran_Mandiri_2026.xlsx")
wb.save(output_filename)
print(f"Successfully generated {output_filename}")
